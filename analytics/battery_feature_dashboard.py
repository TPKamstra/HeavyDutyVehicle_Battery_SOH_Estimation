"""
battery_feature_dashboard.py

Interactive feature dashboard for testday_run BDPS files.

Each testday_run file is a pulse test: battery sits at rest, then receives
discharge pulses (I < 0) and charge pulses (I > 0) at ~10 Hz.

Tabs
────
  Run Inspector      – annotated V/I signal + DCIR per pulse + SoC estimate
  Compare Runs       – overlay multiple runs on a shared time axis
  SOH History        – SOH from discharge_c5 files, linked to next 2 testday_runs
  Degradation Trends – per-feature trend across all testday_runs

Run with:
    panel serve battery_feature_dashboard.py --show
"""

import io
import os
import re
import warnings

import numpy as np
import pandas as pd
import panel as pn
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from scipy import stats as _scipy_stats

import testday_v2_features as v2feat

warnings.filterwarnings("ignore")
pn.extension("plotly", "tabulator")

# ── Config ────────────────────────────────────────────────────────────────────
BASE_DIR       = os.path.dirname(__file__)
LOG_DIR        = os.path.join(BASE_DIR, "LOGBATTEST_Complete")
DCIR_WIN_S     = 3.0    # seconds at pulse start used for DCIR average
OCV_WIN_S      = 2.0    # seconds of prior rest used to estimate V_OCV
PULSE_THRESH_A = 1.0    # |I| > this → pulse active
REST_THRESH_A  = 0.2    # |I| < this → rest
TESTDAY_CUTOFF = pd.Timestamp("2026-01-27")  # ignore testday_run files before this date
SOH_MIN_BYTES  = 20_000  # discharge_c5 files smaller than this are fragments
SOH_MIN_AH     = 5.0     # minimum discharge capacity to count as a valid SOH measurement
SOH_CUTOFF_V   = 11.0    # final voltage above this → discharge did not reach cutoff (incomplete)

_trapz = np.trapezoid if hasattr(np, "trapezoid") else np.trapz

_COLORS = [
    "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
    "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf",
]

# ── OCV–SoC lookup (Ultracell UL18-12, 12V 18Ah VRLA/AGM, 6 cells, 25°C) ─────
# Source: standard VRLA/AGM resting-voltage table; battery must be rested ≥ 1 h
_OCV_V   = np.array([10.50, 11.51, 11.66, 11.81, 11.96, 12.10, 12.20, 12.32, 12.42, 12.50, 12.70])
_SOC_PCT = np.array([   0,    10,    20,    30,    40,    50,    60,    70,    80,    90,   100])


def _v_to_soc(v_ocv: float) -> float:
    return round(float(np.interp(v_ocv, _OCV_V, _SOC_PCT)), 1)


# ── File discovery ─────────────────────────────────────────────────────────────
_TESTDAY_PAT = re.compile(
    r"^testday_run_(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})_bdps\.csv$"
)
_DIS_C5_PAT = re.compile(
    r"^discharge_c5_(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2}-\d{2})_bdps\.csv$"
)
# SoCsweep_OCV12p35V_testday_bdps_YYYY-MM-DD_HH-MM-SS.csv
# Block_01_OCV12p35V_testday_bdps_YYYY-MM-DD_HH-MM-SS.csv
_SOC_SWEEP_PAT = re.compile(
    r"^(?:Block_(\d+)_|SoCsweep_)"
    r"OCV(\d+p\d+)V_testday_bdps_"
    r"(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})\.csv$"
)

# OCV → nominal SoC lookup used by the Pi sweep code (Ultracell UL18-12 at 25 °C)
_SWEEP_OCV_V   = np.array([11.58, 11.80, 12.00, 12.12, 12.20, 12.28, 12.36, 12.46, 12.58, 12.70])
_SWEEP_SOC_PCT = np.array([  10,    20,    30,    40,    50,    60,    70,    80,    90,   100])


def find_runs(log_dir: str) -> list:
    """All testday_run timestamps on or after TESTDAY_CUTOFF, sorted."""
    all_ts = sorted(
        {_TESTDAY_PAT.match(fn).group(1)
         for fn in os.listdir(log_dir) if _TESTDAY_PAT.match(fn)}
    )
    return [ts for ts in all_ts if pd.Timestamp(ts[:10]) >= TESTDAY_CUTOFF]


def find_all_runs(log_dir: str) -> list:
    """All testday_run timestamps regardless of cutoff (used for SOH linking)."""
    return sorted(
        {_TESTDAY_PAT.match(fn).group(1)
         for fn in os.listdir(log_dir) if _TESTDAY_PAT.match(fn)}
    )


def _sweep_ocv_to_soc(ocv: float) -> float:
    return round(float(np.interp(ocv, _SWEEP_OCV_V, _SWEEP_SOC_PCT)), 1)


def _is_v2_schema(path: str) -> bool:
    """True if this file already uses the new event-scripted test-day profile
    (Event_Type column) — those belong exclusively to the Test-Day v2 (Beta)
    tab, not the old threshold-based pulse detection used here."""
    try:
        return "Event_Type" in pd.read_csv(path, nrows=0).columns
    except Exception:
        return False


def find_soc_sweep_runs(log_dir: str) -> list:
    """
    Return list of dicts for SoCsweep / Block testday BDPS files.
    Each dict has: file, path, block (int or None), session (ts of the sweep start),
    ocv_filename, soc_nominal, ts, dt.
    All files with the same filename timestamp share the same `session` key.
    Skips files already using the new v2 event-scripted schema — see
    testday_v2_features.py / the Test-Day v2 (Beta) tab for those.
    """
    results = []
    for fn in os.listdir(log_dir):
        m = _SOC_SWEEP_PAT.match(fn)
        if not m:
            continue
        if _is_v2_schema(os.path.join(log_dir, fn)):
            continue
        block  = int(m.group(1)) if m.group(1) else None
        ocv_v  = float(m.group(2).replace("p", "."))
        ts_str = m.group(3)
        dt     = pd.Timestamp(ts_str[:10] + " " + ts_str[11:].replace("-", ":"))
        results.append({
            "file":         fn,
            "path":         os.path.join(log_dir, fn),
            "block":        block,
            "session":      ts_str,   # groups all files from the same sweep run
            "ocv_filename": ocv_v,
            "soc_nominal":  _sweep_ocv_to_soc(ocv_v),
            "ts":           ts_str,
            "dt":           dt,
        })
    return sorted(results, key=lambda x: (x["dt"], x["block"] if x["block"] is not None else -1, x["ocv_filename"]))


def find_discharge_c5() -> list:
    """
    Return (datetime, filename, filepath) tuples for discharge_c5 BDPS files
    that are large enough to be full (or near-full) discharges.
    """
    results = []
    for fn in os.listdir(LOG_DIR):
        m = _DIS_C5_PAT.match(fn)
        if not m:
            continue
        path = os.path.join(LOG_DIR, fn)
        if os.path.getsize(path) < SOH_MIN_BYTES:
            continue
        dt = pd.Timestamp(f"{m.group(1)} {m.group(2).replace('-', ':')}")
        results.append((dt, fn, path))
    return sorted(results, key=lambda x: x[0])


def _compute_cycle_map() -> dict:
    """Map each testday_run timestamp → cumulative cycle count (unique date × 10)."""
    if not RUNS:
        return {}
    unique_dates = sorted(set(ts[:10] for ts in RUNS))
    date_to_cycle = {d: i * 10 for i, d in enumerate(unique_dates)}
    return {ts: date_to_cycle[ts[:10]] for ts in RUNS}


# ── Data loading ───────────────────────────────────────────────────────────────
def load_run(ts: str) -> pd.DataFrame:
    path = os.path.join(LOG_DIR, f"testday_run_{ts}_bdps.csv")
    if not os.path.exists(path):
        return pd.DataFrame()
    try:
        df = pd.read_csv(path)
        df.columns = [c.strip().lower() for c in df.columns]
        for col in ("elapsed_s", "voltage", "current"):
            df[col] = pd.to_numeric(df[col], errors="coerce")
        return df.dropna(subset=["elapsed_s", "voltage", "current"]).reset_index(drop=True)
    except Exception:
        return pd.DataFrame()


def _load_dis_csv(path: str) -> pd.DataFrame:
    try:
        df = pd.read_csv(path)
        df.columns = [c.strip().lower() for c in df.columns]
        for col in ("elapsed_s", "voltage", "current"):
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        return df.dropna(subset=["current"]).reset_index(drop=True)
    except Exception:
        return pd.DataFrame()


# ── Pulse detection ────────────────────────────────────────────────────────────
def detect_pulses(df: pd.DataFrame) -> list:
    if df.empty:
        return []
    t = df["elapsed_s"].to_numpy()
    v = df["voltage"].to_numpy()
    i = df["current"].to_numpy()
    state       = np.where(i < -PULSE_THRESH_A, "D",
                  np.where(i >  PULSE_THRESH_A, "C", "R"))
    pulses      = []
    pulse_start = None
    pulse_type  = None
    for k, s in enumerate(state):
        if s in ("D", "C") and pulse_start is None:
            pulse_start = k; pulse_type = s
        elif pulse_start is not None and s != pulse_type:
            if k - 1 > pulse_start:
                p = _pulse_features(t, v, i, pulse_start, k - 1, pulse_type)
                if p:
                    pulses.append(p)
            pulse_start = None; pulse_type = None
            if s in ("D", "C"):
                pulse_start = k; pulse_type = s
    if pulse_start is not None and len(state) - 1 > pulse_start:
        p = _pulse_features(t, v, i, pulse_start, len(state) - 1, pulse_type)
        if p:
            pulses.append(p)
    return pulses


def _pulse_features(t, v, i, i0, i1, ptype) -> dict | None:
    duration = float(t[i1] - t[i0])
    if duration < 0.2:
        return None
    t0        = float(t[i0])
    t_end     = float(t[i1])

    # OCV: mean voltage in rest window before pulse
    rest_mask = (t >= t0 - OCV_WIN_S) & (t < t0) & (np.abs(i) < REST_THRESH_A)
    v_ocv     = float(np.mean(v[rest_mask])) if rest_mask.any() else float(v[max(0, i0 - 1)])

    # DCIR: first DCIR_WIN_S seconds of pulse
    dcir_mask = (t >= t0) & (t <= t0 + DCIR_WIN_S)
    if dcir_mask.sum() < 2:
        dcir_mask = np.zeros(len(t), bool); dcir_mask[i0 : i1 + 1] = True
    v_pulse   = float(np.mean(v[dcir_mask]))
    i_pulse   = float(np.mean(i[dcir_mask]))
    di        = abs(i_pulse)
    dcir      = round(abs(v_ocv - v_pulse) / di * 1000, 1) if di > 0.1 else None

    # End-of-pulse voltage: last 2 s of pulse
    eop_mask  = (t >= max(t0, t_end - 2.0)) & (t <= t_end)
    v_eop     = round(float(np.mean(v[eop_mask])), 3) if eop_mask.sum() > 1 else round(float(v[i1]), 3)

    # Voltage recovery: first 2 s of rest immediately after pulse
    recover_mask = (t > t_end) & (t <= t_end + 2.0) & (np.abs(i) < REST_THRESH_A)
    v_recover    = round(float(np.mean(v[recover_mask])), 3) if recover_mask.any() else None

    # Charge throughput
    seg_t     = t[i0 : i1 + 1]
    seg_i     = i[i0 : i1 + 1]
    dt_seg    = np.clip(np.diff(seg_t, prepend=seg_t[0]), 0.0, 10.0)
    dt_seg[0] = 0.0
    q_ah      = float(np.nansum(np.abs(seg_i) * dt_seg) / 3600.0)
    i_peak    = float(np.min(i[i0 : i1 + 1])) if ptype == "D" else float(np.max(i[i0 : i1 + 1]))

    # Charge-acceptance features (charge pulses only)
    dvdt_mVs   = None
    v_chg_peak = None
    if ptype == "C" and duration >= 5.0:
        v_chg_peak = round(float(np.max(v[i0 : i1 + 1])), 3)
        dvdt_mVs   = round((v_eop - v_pulse) / max(duration, 1.0) * 1000, 3)

    return {
        "type": "DIS" if ptype == "D" else "CHG",
        "idx_start": int(i0), "idx_end": int(i1),
        "t_start": round(t0, 2), "t_end": round(t_end, 2),
        "duration_s": round(duration, 1), "i_peak": round(i_peak, 2),
        "v_ocv": round(v_ocv, 3), "v_pulse": round(v_pulse, 3),
        "dv": round(abs(v_ocv - v_pulse), 3),
        "dcir_mohm": dcir, "q_ah": round(q_ah, 4),
        "soc": _v_to_soc(v_ocv),
        "v_eop": v_eop, "v_recover": v_recover,
        "dvdt_mVs": dvdt_mVs, "v_chg_peak": v_chg_peak,
    }


# ── Session summary ────────────────────────────────────────────────────────────
def session_summary(ts: str) -> dict:
    df  = load_run(ts)
    row = {"session": ts, "date": ts[:10]}
    if df.empty:
        return row
    pulses = detect_pulses(df)
    dis    = [p for p in pulses if p["type"] == "DIS"]
    chg    = [p for p in pulses if p["type"] == "CHG"]

    def _mean(lst, key):
        vals = [p[key] for p in lst if p.get(key) is not None]
        return round(float(np.mean(vals)), 2) if vals else None

    i_arr       = df["current"].to_numpy()
    v_arr       = df["voltage"].to_numpy()
    rest        = np.abs(i_arr) < REST_THRESH_A
    v_ocv_start = round(float(np.mean(v_arr[rest][:20])), 3) if rest.any() else None

    q_dis = round(sum(p["q_ah"] for p in dis), 4)
    q_chg = round(sum(p["q_ah"] for p in chg), 4)
    temp_c, hum_pct = _DATE_TEMP_CACHE.get(ts[:10], (None, None))
    row.update({
        "n_dis_pulses":       len(dis),
        "n_chg_pulses":       len(chg),
        "DCIR_dis [mΩ]":     _mean(dis, "dcir_mohm"),
        "DCIR_chg [mΩ]":     _mean(chg, "dcir_mohm"),
        "V_OCV [V]":          v_ocv_start,
        "SoC_start [%]":      _v_to_soc(v_ocv_start) if v_ocv_start is not None else None,
        "Q_dis [Ah]":         q_dis,
        "Q_chg [Ah]":         q_chg,
        "I_dis_peak [A]":     _mean(dis, "i_peak"),
        "dur_dis_mean [s]":   _mean(dis, "duration_s"),
        "V_eop_dis [V]":      _mean(dis, "v_eop"),
        "V_recover_dis [V]":  _mean(dis, "v_recover"),
        "CA_dVdt [mV/s]":     _mean(chg, "dvdt_mVs"),
        "V_chg_peak [V]":     _mean(chg, "v_chg_peak"),
        "Eta_c [%]":          round(q_chg / q_dis * 100, 1) if q_dis > 0 else None,
        "Temp [°C]":          temp_c,
        "Humidity [%]":       hum_pct,
    })
    return row


_RUN_SUMMARIES: dict = {}


def _run_summary_cached(ts: str) -> dict:
    if ts not in _RUN_SUMMARIES:
        _RUN_SUMMARIES[ts] = session_summary(ts)
    return _RUN_SUMMARIES[ts]


def _read_sensor_temp(bdps_path: str) -> tuple:
    """
    Read the matching _sensor_ file for a BDPS file and return
    (mean_temperature_C, mean_humidity_pct). Returns (None, None) if not found.
    """
    sensor_path = bdps_path.replace("_bdps_", "_sensor_")
    if not os.path.exists(sensor_path):
        return None, None
    try:
        df = pd.read_csv(sensor_path, nrows=200)
        df.columns = [c.strip().lower() for c in df.columns]
        temp = round(float(df["temperature"].dropna().mean()), 2) if "temperature" in df.columns else None
        hum  = round(float(df["humidity"].dropna().mean()), 1)    if "humidity"    in df.columns else None
        return temp, hum
    except Exception:
        return None, None


def sweep_session_summary(info: dict) -> dict:
    """Compute features from a SoCsweep / Block testday BDPS file."""
    df = _load_dis_csv(info["path"])
    row = {
        "file":              info["file"],
        "dt":                info["dt"],
        "date":              info["dt"].date().isoformat(),
        "block":             info["block"],
        "session":           info["session"],   # groups files from the same sweep
        "ocv_filename [V]":  info["ocv_filename"],
        "soc_nominal [%]":   info["soc_nominal"],
    }
    if df.empty or "elapsed_s" not in df.columns:
        return row
    for col in ("elapsed_s", "voltage", "current"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(subset=["elapsed_s", "voltage", "current"]).reset_index(drop=True)
    if len(df) < 50:
        return row

    pulses = detect_pulses(df)
    dis    = [p for p in pulses if p["type"] == "DIS"]
    chg    = [p for p in pulses if p["type"] == "CHG"]

    def _mean(lst, key):
        vals = [p[key] for p in lst if p.get(key) is not None]
        return round(float(np.mean(vals)), 2) if vals else None

    t_arr    = df["elapsed_s"].to_numpy()
    v_arr    = df["voltage"].to_numpy()
    i_arr    = df["current"].to_numpy()
    dt_s     = np.clip(np.diff(t_arr, prepend=t_arr[0]), 0.0, 10.0)
    dt_s[0]  = 0.0
    dis_mask = i_arr < -PULSE_THRESH_A

    q_dis      = round(sum(p["q_ah"] for p in dis), 4)
    q_chg      = round(sum(p["q_ah"] for p in chg), 4)
    v_min      = round(float(np.min(v_arr[dis_mask])), 3)       if dis_mask.any() else None
    v_mean_dis = round(float(np.mean(v_arr[dis_mask])), 3)      if dis_mask.any() else None
    i_peak     = round(float(np.min(i_arr[dis_mask])), 2)       if dis_mask.any() else None
    e_dis      = round(
        float(np.nansum(np.abs(i_arr[dis_mask]) * v_arr[dis_mask] * dt_s[dis_mask]) / 3600.0), 3
    )                                                             if dis_mask.any() else None

    temp_c, hum_pct = _read_sensor_temp(info["path"])

    row.update({
        "n_dis_pulses":      len(dis),
        "n_chg_pulses":      len(chg),
        "DCIR_dis [mΩ]":    _mean(dis, "dcir_mohm"),
        "V_min [V]":         v_min,
        "V_mean_dis [V]":    v_mean_dis,
        "I_peak_dis [A]":    i_peak,
        "Q_dis [Ah]":        q_dis,
        "Q_chg [Ah]":        q_chg,
        "E_dis [Wh]":        e_dis,
        "V_eop_dis [V]":     _mean(dis, "v_eop"),
        "V_recover_dis [V]": _mean(dis, "v_recover"),
        "CA_dVdt [mV/s]":   _mean(chg, "dvdt_mVs"),
        "Temp [°C]":         temp_c,
        "Humidity [%]":      hum_pct,
    })
    return row


# ── SOH computation ────────────────────────────────────────────────────────────
def _capacity_from_dis(path: str) -> tuple:
    """
    Compute discharge capacity [Ah] and final voltage [V] from a discharge_c5 BDPS file.
    Returns (capacity_ah, v_final) where capacity_ah is None if below SOH_MIN_AH.
    """
    df = _load_dis_csv(path)
    if df.empty or "elapsed_s" not in df.columns or len(df) < 50:
        return None, None
    t  = df["elapsed_s"].to_numpy(float)
    i  = df["current"].to_numpy(float)
    v  = df["voltage"].to_numpy(float) if "voltage" in df.columns else np.full_like(i, np.nan)
    dt = np.diff(t, prepend=t[0])
    dt[0] = np.nan
    dt = np.where(np.isfinite(dt) & (dt >= 0) & (dt <= 10.0), dt, np.nan)
    i_fin = i[np.isfinite(i)]
    if len(i_fin) == 0:
        return None, None
    # detect discharge direction
    mask = (i < -0.5) if np.mean(i_fin < -0.5) >= np.mean(i_fin > 0.5) else (i > 0.5)
    mask = mask & np.isfinite(i) & np.isfinite(dt)
    if not mask.any():
        return None, None
    ah = float(abs(np.nansum(i[mask] * dt[mask]) / 3600.0))
    v_final = float(v[np.isfinite(v)][-1]) if np.any(np.isfinite(v)) else None
    if ah < SOH_MIN_AH:
        return None, v_final
    return ah, v_final


def compute_soh_series(nominal_ah: float = 18.0, baseline: str = "first") -> pd.DataFrame:
    """
    Build SOH timeseries from all large discharge_c5 BDPS files.
    baseline = "first"   → SOH relative to first valid file
             = "nominal" → SOH relative to nominal_ah
    Adds 'complete' column: True when discharge reached the cutoff voltage.
    """
    rows = []
    for dt, fn, path in find_discharge_c5():
        cap, v_final = _capacity_from_dis(path)
        if cap is None:
            continue
        complete = (v_final is None) or (v_final <= SOH_CUTOFF_V)
        temp_c, hum_pct = _DATE_TEMP_CACHE.get(dt.date().isoformat(), (None, None))
        rows.append({"datetime": dt, "date": dt.date().isoformat(),
                     "file": fn, "capacity_ah": cap,
                     "v_final": v_final, "complete": complete,
                     "temp_c": temp_c, "humidity_pct": hum_pct})
    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows).sort_values("datetime").reset_index(drop=True)
    if baseline == "first":
        b = float(out["capacity_ah"].iloc[0])
    else:
        b = float(nominal_ah)
    out["soh_pct"] = (out["capacity_ah"] / b * 100).round(1)
    return out


def link_soh_to_runs(soh_df: pd.DataFrame) -> pd.DataFrame:
    """
    For each SOH row, find the 2 testday_run files immediately afterwards
    (from the filtered RUNS list), and attach their DCIR.
    """
    if soh_df.empty or not _ALL_RUNS:
        return soh_df
    run_times = sorted(
        [(ts, pd.Timestamp(ts[:10] + " " + ts[11:].replace("-", ":"))) for ts in _ALL_RUNS],
        key=lambda x: x[1],
    )
    out = soh_df.copy()
    for c in ("run_1", "run_1_dt", "dcir_1", "run_1_cycle",
              "run_2", "run_2_dt", "dcir_2", "run_2_cycle",
              "soh_cycle"):
        out[c] = None

    for idx, row in out.iterrows():
        dt    = pd.Timestamp(row["datetime"])
        after = [(ts, t) for ts, t in run_times if t > dt]
        for j, (key_ts, key_dt, key_dcir, key_cyc) in enumerate([
            ("run_1", "run_1_dt", "dcir_1", "run_1_cycle"),
            ("run_2", "run_2_dt", "dcir_2", "run_2_cycle"),
        ]):
            if j < len(after):
                r_ts, r_dt = after[j]
                s = _run_summary_cached(r_ts)
                out.at[idx, key_ts]   = r_ts
                out.at[idx, key_dt]   = r_dt
                out.at[idx, key_dcir] = s.get("DCIR_dis [mΩ]")
                out.at[idx, key_cyc]  = CYCLE_MAP.get(r_ts)
        # SOH cycle = cycle of first following run (will be filled in below for entries without one)
        out.at[idx, "soh_cycle"] = out.at[idx, "run_1_cycle"]

    # For SOH entries with no linked testday_run after them, assign sequential cycles
    # starting at max_known_cycle + 10 (10 degradation cycles between each SOH measurement).
    max_known = max(CYCLE_MAP.values()) if CYCLE_MAP else 0
    no_cycle = out[out["soh_cycle"].isna()].sort_values("datetime")
    for rank, idx in enumerate(no_cycle.index, start=1):
        out.at[idx, "soh_cycle"] = max_known + rank * 10

    return out


def link_soh_to_sweeps(soh_df: pd.DataFrame) -> pd.DataFrame:
    """
    For each SOH row, find the most recent SoCsweep session that precedes it
    and attach the session key plus the per-OCV DCIR profile from _SWEEP_DF.

    Adds columns:
        sweep_session      – ts string of the linked sweep session (e.g. "2026-06-18_14-32-54")
        sweep_session_dt   – datetime of that session
        sweep_dcir_mean    – mean DCIR [mΩ] across all OCV runs in that session
        sweep_n_runs       – number of sweep runs linked
    """
    if soh_df.empty or not SOC_SWEEP_RUNS:
        return soh_df

    # Collect unique sessions sorted chronologically
    seen: dict = {}
    for info in SOC_SWEEP_RUNS:
        ts = info["session"]
        if ts not in seen:
            seen[ts] = info["dt"]
    session_list = sorted(seen.items(), key=lambda x: x[1])   # [(ts, dt), …]

    out = soh_df.copy()
    for col in ("sweep_session", "sweep_session_dt", "sweep_dcir_mean", "sweep_n_runs"):
        out[col] = None

    for idx, row in out.iterrows():
        soh_dt = pd.Timestamp(row["datetime"])
        before = [(ts, dt) for ts, dt in session_list if dt < soh_dt]
        if not before:
            continue
        sess_ts, sess_dt = before[-1]
        out.at[idx, "sweep_session"]    = sess_ts
        out.at[idx, "sweep_session_dt"] = sess_dt

        # Pull DCIR from _SWEEP_DF if features have been computed
        if not _SWEEP_DF.empty and "session" in _SWEEP_DF.columns and "DCIR_dis [mΩ]" in _SWEEP_DF.columns:
            sess_rows  = _SWEEP_DF[_SWEEP_DF["session"] == sess_ts]
            dcir_vals  = sess_rows["DCIR_dis [mΩ]"].dropna()
            if not dcir_vals.empty:
                out.at[idx, "sweep_dcir_mean"] = round(float(dcir_vals.mean()), 1)
                out.at[idx, "sweep_n_runs"]    = int(len(sess_rows))

    return out


def _annotate_sweeps_with_soh(sweep_df: pd.DataFrame, soh_df: pd.DataFrame) -> pd.DataFrame:
    """
    Add 'linked_soh_pct' column to _SWEEP_DF rows so each session line
    in build_sweep_fig() can be labelled with the subsequent SOH value.
    """
    if sweep_df.empty or soh_df.empty or "sweep_session" not in soh_df.columns:
        sweep_df["linked_soh_pct"] = None
        return sweep_df
    # Map session ts → SOH %
    soh_map = (
        soh_df[soh_df["sweep_session"].notna()]
        .set_index("sweep_session")["soh_pct"]
        .to_dict()
    )
    sweep_df = sweep_df.copy()
    sweep_df["linked_soh_pct"] = sweep_df["session"].map(soh_map)
    return sweep_df


# ── Plotly config helper ───────────────────────────────────────────────────────
def _plotly_cfg(stem: str) -> dict:
    return {
        "displaylogo": False, "responsive": True,
        "toImageButtonOptions": {
            "format": "png", "filename": stem,
            "height": 1200, "width": 2000, "scale": 2,
        },
    }


# ── Signal figure ──────────────────────────────────────────────────────────────
def build_signal_fig(ts: str) -> go.Figure:
    df = load_run(ts)
    if df.empty:
        return go.Figure().update_layout(title="No data for this run")
    t  = df["elapsed_s"].to_numpy()
    v  = df["voltage"].to_numpy()
    i  = df["current"].to_numpy()
    pulses  = detect_pulses(df)
    rest    = np.abs(i) < REST_THRESH_A
    v_start = float(np.mean(v[rest][:20])) if rest.any() else float(v[0])
    soc_str = f"  ·  SoC_start ≈ {_v_to_soc(v_start):.0f} %  (V_OCV = {v_start:.3f} V)"

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(go.Scatter(x=t, y=v, name="Voltage [V]", mode="lines",
                             line=dict(color="#1f77b4", width=1.8)), secondary_y=False)
    fig.add_trace(go.Scatter(x=t, y=i, name="Current [A]", mode="lines",
                             line=dict(color="#999", width=1.0), opacity=0.65), secondary_y=True)
    for p in pulses:
        color = "rgba(255,127,14,0.14)" if p["type"] == "DIS" else "rgba(31,119,180,0.12)"
        fig.add_vrect(x0=p["t_start"], x1=p["t_end"],
                      fillcolor=color, opacity=1.0, layer="below", line_width=0)
        if p.get("dcir_mohm") is not None:
            ann_col = "#d62728" if p["type"] == "DIS" else "#2ca02c"
            fig.add_annotation(
                x=(p["t_start"] + p["t_end"]) / 2,
                y=p["v_ocv"] + (0.2 if p["type"] == "DIS" else -0.2),
                text=f"{p['dcir_mohm']} mΩ", showarrow=False,
                font=dict(size=9, color=ann_col),
                bgcolor="rgba(255,255,255,0.85)",
            )
    fig.update_layout(
        title=f"Pulse test — {ts[:10]}  {ts[11:].replace('-', ':')}{soc_str}",
        xaxis_title="Elapsed [s]",
        margin=dict(l=50, r=50, t=65, b=40), height=430,
        legend=dict(orientation="h", y=1.13),
    )
    fig.update_yaxes(title_text="Voltage [V]", secondary_y=False)
    fig.update_yaxes(title_text="Current [A]", secondary_y=True)
    return fig


def build_pulse_table(ts: str) -> pd.DataFrame:
    df = load_run(ts)
    if df.empty:
        return pd.DataFrame()
    pulses = detect_pulses(df)
    if not pulses:
        return pd.DataFrame()
    return pd.DataFrame([{
        "Pulse #": k + 1, "Type": p["type"],
        "t_start [s]": p["t_start"], "Duration [s]": p["duration_s"],
        "I_peak [A]": p["i_peak"], "V_OCV [V]": p["v_ocv"],
        "SoC [%]": p["soc"], "V_pulse [V]": p["v_pulse"],
        "V_eop [V]": p["v_eop"], "V_recover [V]": p.get("v_recover"),
        "ΔV [V]": p["dv"], "DCIR [mΩ]": p["dcir_mohm"], "Q [Ah]": p["q_ah"],
        "dV/dt [mV/s]": p.get("dvdt_mVs"),
    } for k, p in enumerate(pulses)])


# ── Compare figure ─────────────────────────────────────────────────────────────
def _load_csv_by_path(path: str) -> pd.DataFrame:
    try:
        df = pd.read_csv(path)
        df.columns = [c.strip().lower() for c in df.columns]
        for col in ("elapsed_s", "voltage", "current"):
            df[col] = pd.to_numeric(df[col], errors="coerce")
        return df.dropna(subset=["elapsed_s", "voltage", "current"]).reset_index(drop=True)
    except Exception:
        return pd.DataFrame()


def build_compare_fig(run_list: list, sweep_paths: list | None = None) -> go.Figure:
    fig = make_subplots(rows=2, cols=1, shared_xaxes=True,
                        row_heights=[0.6, 0.4], vertical_spacing=0.06)
    sweep_paths = sweep_paths or []
    if not run_list and not sweep_paths:
        return fig.update_layout(title="Select at least one run to compare")

    traces = []  # (label, group_key, df) — collected before assigning colours

    for ts in run_list:
        df = load_run(ts)
        if df.empty:
            continue
        i = df["current"].to_numpy()
        v = df["voltage"].to_numpy()
        rest    = np.abs(i) < REST_THRESH_A
        v_start = float(np.mean(v[rest][:20])) if rest.any() else float(v[0])
        label   = ts[:10] + " " + ts[11:].replace("-", ":") + f"  SoC≈{_v_to_soc(v_start):.0f}%"
        traces.append((label, ts, df))

    sweep_info_by_path = {info["path"]: info for info in SOC_SWEEP_RUNS}
    for path in sweep_paths:
        df = _load_csv_by_path(path)
        if df.empty:
            continue
        info = sweep_info_by_path.get(path, {})
        sess_date = info.get("ts", "")[:10]
        ocv_v     = info.get("ocv_filename", "?")
        soc_pct   = info.get("soc_nominal", "?")
        label     = f"{sess_date}  OCV {ocv_v:.2f}V  SoC≈{soc_pct:.0f}%  [sweep]"
        traces.append((label, path, df))

    for k, (label, group_key, df) in enumerate(traces):
        color = _COLORS[k % len(_COLORS)]
        t     = df["elapsed_s"].to_numpy()
        v     = df["voltage"].to_numpy()
        i     = df["current"].to_numpy()
        step  = max(1, len(t) // 3000)
        fig.add_trace(go.Scatter(x=t[::step], y=v[::step], name=label,
                                 mode="lines", line=dict(color=color, width=1.5),
                                 legendgroup=group_key), row=1, col=1)
        fig.add_trace(go.Scatter(x=t[::step], y=i[::step], name=label + " (I)",
                                 mode="lines", line=dict(color=color, width=1.0, dash="dot"),
                                 opacity=0.7, legendgroup=group_key, showlegend=False), row=2, col=1)

    fig.update_yaxes(title_text="Voltage [V]", row=1, col=1)
    fig.update_yaxes(title_text="Current [A]", row=2, col=1)
    fig.update_xaxes(title_text="Elapsed [s]", row=2, col=1)
    fig.update_layout(title="Run comparison — voltage (top) · current (bottom)",
                      margin=dict(l=50, r=50, t=65, b=40), height=550,
                      legend=dict(orientation="v", x=1.02, y=1.0))
    return fig


# ── SOH figure ─────────────────────────────────────────────────────────────────
def build_soh_fig(soh_df: pd.DataFrame, x_mode: str = "Date") -> go.Figure:
    if soh_df.empty:
        return go.Figure().update_layout(title="No SOH data — click Compute SOH")

    has_sweep = "sweep_dcir_mean" in soh_df.columns and soh_df["sweep_dcir_mean"].notna().any()
    has_temp  = "temp_c" in soh_df.columns and soh_df["temp_c"].notna().any()

    if has_sweep and has_temp:
        n_rows, heights = 4, [0.38, 0.24, 0.20, 0.18]
        subtitles = (
            "SOH [%]  —  from discharge_c5 files",
            "DCIR [mΩ]  —  from linked testday_run files",
            "Mean sweep DCIR [mΩ]  —  from linked SoCsweep session",
            "Temperature [°C]  —  ambient during discharge test",
        )
    elif has_sweep:
        n_rows, heights = 3, [0.45, 0.30, 0.25]
        subtitles = (
            "SOH [%]  —  from discharge_c5 files",
            "DCIR [mΩ]  —  from linked testday_run files",
            "Mean sweep DCIR [mΩ]  —  from linked SoCsweep session",
        )
    elif has_temp:
        n_rows, heights = 3, [0.45, 0.33, 0.22]
        subtitles = (
            "SOH [%]  —  from discharge_c5 files",
            "DCIR [mΩ]  —  from linked testday_run files",
            "Temperature [°C]  —  ambient during discharge test",
        )
    else:
        n_rows, heights = 2, [0.55, 0.45]
        subtitles = (
            "SOH [%]  —  from discharge_c5 files",
            "DCIR [mΩ]  —  from linked testday_run files",
        )

    fig = make_subplots(
        rows=n_rows, cols=1, shared_xaxes=True,
        row_heights=heights, vertical_spacing=0.07,
        subplot_titles=subtitles,
    )
    use_cycles = x_mode == "Cycle count"

    # ── SOH line
    if use_cycles:
        soh_mask  = soh_df["soh_cycle"].notna()
        sub_soh   = soh_df[soh_mask]
        soh_x     = sub_soh["soh_cycle"].astype(float)
        soh_htmpl = (
            "Cycle %{x:.0f}<br>"
            "SOH = %{y:.1f} %<br>"
            "Cap = %{customdata[0]:.2f} Ah<br>"
            "%{customdata[1]}<extra></extra>"
        )
        x_label = "Cycle count"
    else:
        sub_soh   = soh_df
        soh_x     = pd.to_datetime(soh_df["datetime"])
        soh_htmpl = (
            "%{x|%Y-%m-%d %H:%M}<br>"
            "SOH = %{y:.1f} %<br>"
            "Cap = %{customdata[0]:.2f} Ah<br>"
            "%{customdata[1]}<extra></extra>"
        )
        x_label = "Date"

    # Split complete vs incomplete discharges
    has_complete_col = "complete" in sub_soh.columns
    comp_mask = sub_soh["complete"].astype(bool) if has_complete_col else pd.Series(True, index=sub_soh.index)
    inc_mask  = ~comp_mask

    # Complete discharges — solid filled circle
    if comp_mask.any():
        s = sub_soh[comp_mask]
        fig.add_trace(go.Scatter(
            x=soh_x[comp_mask], y=s["soh_pct"],
            mode="lines+markers", name="SOH [%]",
            line=dict(color="#1f77b4", width=2.0),
            marker=dict(size=9, color="#1f77b4"),
            customdata=np.stack([s["capacity_ah"], s["file"]], axis=1),
            hovertemplate=soh_htmpl,
        ), row=1, col=1)

    # Incomplete discharges — open circle with warning colour
    if inc_mask.any():
        s = sub_soh[inc_mask]
        vf = s["v_final"].round(2) if "v_final" in s.columns else pd.Series(["?"] * len(s))
        inc_cdata = np.stack([s["capacity_ah"], s["file"],
                              vf.fillna("?").astype(str)], axis=1)
        fig.add_trace(go.Scatter(
            x=soh_x[inc_mask], y=s["soh_pct"],
            mode="markers", name="SOH [%] — incomplete",
            marker=dict(size=12, color="#ff7f0e", symbol="circle-open", line=dict(width=2.5)),
            customdata=inc_cdata,
            hovertemplate=(
                soh_htmpl.replace(
                    "%{customdata[1]}<extra></extra>",
                    "%{customdata[1]}<br>V_final = %{customdata[2]} V  ⚠ discharge cut short<extra></extra>",
                )
            ),
        ), row=1, col=1)

    # ── Linked testday_run DCIR (middle panel)
    for j, (dcir_col, dt_col, cyc_col, color, sym, name) in enumerate([
        ("dcir_1", "run_1_dt", "run_1_cycle", "#ff7f0e", "circle",  "DCIR — linked run 1"),
        ("dcir_2", "run_2_dt", "run_2_cycle", "#d62728", "diamond", "DCIR — linked run 2"),
    ]):
        mask = soh_df[dt_col].notna() & soh_df[dcir_col].notna()
        if use_cycles:
            mask = mask & soh_df[cyc_col].notna()
        if not mask.any():
            continue
        sub     = soh_df[mask]
        run_key = "run_1" if j == 0 else "run_2"
        if use_cycles:
            dcir_x     = sub[cyc_col].astype(float)
            dcir_htmpl = (
                "Cycle %{x:.0f}<br>"
                "DCIR = %{y:.1f} mΩ<br>"
                "Linked SOH = %{customdata[0]:.1f} %<br>"
                "%{customdata[1]}<extra></extra>"
            )
        else:
            dcir_x     = pd.to_datetime(sub[dt_col].tolist())
            dcir_htmpl = (
                "%{x|%Y-%m-%d %H:%M}<br>"
                "DCIR = %{y:.1f} mΩ<br>"
                "Linked SOH = %{customdata[0]:.1f} %<br>"
                "%{customdata[1]}<extra></extra>"
            )
        fig.add_trace(go.Scatter(
            x=dcir_x, y=sub[dcir_col].astype(float),
            mode="markers", name=name,
            marker=dict(size=11, symbol=sym, color=color,
                        line=dict(width=1, color="white")),
            customdata=np.stack([sub["soh_pct"], sub[run_key].astype(str)], axis=1),
            hovertemplate=dcir_htmpl,
        ), row=2, col=1)

    # ── Shaded bands SOH → linked testday_run
    for _, row in soh_df.iterrows():
        if use_cycles:
            soh_cyc = row["soh_cycle"]
            if soh_cyc is None:
                continue
            for cyc_col in ("run_1_cycle", "run_2_cycle"):
                r_cyc = row[cyc_col]
                if r_cyc is not None and r_cyc != soh_cyc:
                    fig.add_vrect(x0=min(soh_cyc, r_cyc), x1=max(soh_cyc, r_cyc),
                                  fillcolor="rgba(100,100,100,0.05)",
                                  opacity=1.0, layer="below", line_width=0, row=1, col=1)
        else:
            soh_dt = pd.Timestamp(row["datetime"])
            for dt_col in ("run_1_dt", "run_2_dt"):
                if row[dt_col] is not None:
                    fig.add_vrect(x0=soh_dt, x1=pd.Timestamp(row[dt_col]),
                                  fillcolor="rgba(100,100,100,0.05)",
                                  opacity=1.0, layer="below", line_width=0, row=1, col=1)

    # ── Mean sweep DCIR (bottom panel, only when sweep features computed)
    if has_sweep:
        sw_mask = soh_df["sweep_dcir_mean"].notna() & soh_df["sweep_session_dt"].notna()
        if sw_mask.any():
            sw_sub = soh_df[sw_mask]
            if use_cycles:
                sw_x     = sw_sub["soh_cycle"].astype(float) if "soh_cycle" in sw_sub else None
            else:
                sw_x     = pd.to_datetime(sw_sub["sweep_session_dt"].tolist())
            if sw_x is not None:
                n_vals = sw_sub["sweep_n_runs"].fillna("?").astype(str)
                fig.add_trace(go.Scatter(
                    x=sw_x, y=sw_sub["sweep_dcir_mean"].astype(float),
                    mode="markers+text", name="Sweep DCIR (mean)",
                    marker=dict(size=12, symbol="star", color="#9467bd",
                                line=dict(width=1, color="white")),
                    text=["n=" + v for v in n_vals],
                    textposition="top center",
                    customdata=np.stack([sw_sub["soh_pct"], sw_sub["sweep_session"].astype(str)], axis=1),
                    hovertemplate=(
                        "Sweep session: %{customdata[1]}<br>"
                        "Mean DCIR = %{y:.1f} mΩ  (n=%{text})<br>"
                        "Linked SOH = %{customdata[0]:.1f} %<extra></extra>"
                    ),
                ), row=3, col=1)

    # ── Temperature panel
    if has_temp:
        temp_row = 4 if has_sweep else 3
        temp_mask = soh_df["temp_c"].notna()
        if use_cycles:
            temp_mask = temp_mask & soh_df["soh_cycle"].notna()
            tx = soh_df.loc[temp_mask, "soh_cycle"].astype(float)
        else:
            tx = pd.to_datetime(soh_df.loc[temp_mask, "datetime"].tolist())
        ts = soh_df.loc[temp_mask, "temp_c"].astype(float)
        hs = soh_df.loc[temp_mask, "humidity_pct"] if "humidity_pct" in soh_df.columns else None
        cdata = (
            np.stack([ts, hs.fillna(float("nan")).astype(float)], axis=1)
            if hs is not None and hs.notna().any()
            else ts.values.reshape(-1, 1)
        )
        htmpl = (
            "%{customdata[0]:.1f} °C  (humidity %{customdata[1]:.0f} %)<extra></extra>"
            if cdata.shape[1] == 2
            else "%{customdata[0]:.1f} °C<extra></extra>"
        )
        fig.add_trace(go.Scatter(
            x=tx, y=ts,
            mode="markers+lines", name="Temp [°C]",
            line=dict(color="#e377c2", width=1.5, dash="dot"),
            marker=dict(size=8, color="#e377c2"),
            customdata=cdata, hovertemplate=htmpl,
        ), row=temp_row, col=1)

    last_row = n_rows
    fig.update_yaxes(title_text="SOH [%]",          row=1,        col=1, range=[0, 115])
    fig.update_yaxes(title_text="DCIR [mΩ]",        row=2,        col=1)
    if has_sweep:
        fig.update_yaxes(title_text="Sweep DCIR [mΩ]", row=3,     col=1)
    if has_temp:
        fig.update_yaxes(title_text="Temp [°C]", row=(4 if has_sweep else 3), col=1)
    fig.update_xaxes(title_text=x_label,             row=last_row, col=1)
    title_parts = ["SOH history"]
    if has_sweep: title_parts.append("sweep DCIR")
    if has_temp:  title_parts.append("temperature")
    fig.update_layout(
        title=" · ".join(title_parts),
        height=700 if (has_sweep and has_temp) else (650 if (has_sweep or has_temp) else 580),
        margin=dict(l=55, r=55, t=90, b=40),
        legend=dict(orientation="h", y=1.08),
    )
    return fig


def build_soh_table(soh_df: pd.DataFrame) -> pd.DataFrame:
    if soh_df.empty:
        return pd.DataFrame()
    rows = []
    for _, r in soh_df.iterrows():
        complete = bool(r["complete"]) if "complete" in r.index else True
        temp_c = r["temp_c"] if "temp_c" in r.index and r["temp_c"] is not None else None
        row_d = {
            "Discharge date":  str(r["datetime"])[:16],
            "Complete":        "Yes" if complete else "No (cut short)",
            "V_final [V]":     round(r["v_final"], 3) if ("v_final" in r.index and r["v_final"] is not None) else "—",
            "Capacity [Ah]":   round(r["capacity_ah"], 2),
            "SOH [%]":         r["soh_pct"],
            "Temp [°C]":       round(temp_c, 1) if temp_c is not None else "—",
            "Linked run 1":    str(r["run_1_dt"])[:16] if r["run_1_dt"] is not None else "—",
            "DCIR_1 [mΩ]":    r["dcir_1"],
            "Linked run 2":    str(r["run_2_dt"])[:16] if r["run_2_dt"] is not None else "—",
            "DCIR_2 [mΩ]":    r["dcir_2"],
        }
        if "sweep_session" in r.index:
            row_d["Sweep session"]     = r["sweep_session"]   if r["sweep_session"]   is not None else "—"
            row_d["Sweep DCIR (mean)"] = r["sweep_dcir_mean"] if r["sweep_dcir_mean"] is not None else "—"
            row_d["Sweep n runs"]      = r["sweep_n_runs"]    if r["sweep_n_runs"]    is not None else "—"
        rows.append(row_d)
    return pd.DataFrame(rows)


# ── Trend figure ───────────────────────────────────────────────────────────────
def build_trend_fig(feature: str, x_mode: str = "Date") -> go.Figure:
    global _TREND_DF
    fig = go.Figure()
    if _TREND_DF.empty or feature not in _TREND_DF.columns:
        return fig.update_layout(title="Click 'Compute trends' first")
    df = _TREND_DF.dropna(subset=[feature]).copy()
    if df.empty:
        return fig.update_layout(title=f"No data for {feature}")
    df = df.sort_values("date")

    if x_mode == "Cycle count":
        df["_cycle"] = df["session"].map(CYCLE_MAP)
        df = df.dropna(subset=["_cycle"])
        if df.empty:
            return fig.update_layout(title=f"No cycle data for {feature}")
        t_cal    = df["_cycle"].to_numpy(float)
        x_vals   = t_cal
        x_label  = "Cycle count"
        cb_title = "Cycles"
    else:
        dates    = pd.to_datetime(df["date"])
        t_cal    = (dates - dates.min()).dt.days.to_numpy()
        x_vals   = dates
        x_label  = "Date"
        cb_title = "Days"

    y_vals = df[feature].to_numpy(float)

    # IQR-based outlier detection
    q1, q3  = np.percentile(y_vals, 25), np.percentile(y_vals, 75)
    iqr     = q3 - q1
    out_mask = (y_vals < q1 - 1.5 * iqr) | (y_vals > q3 + 1.5 * iqr)
    in_mask  = ~out_mask

    # Outlier markers (faded ✕) — plotted first so they sit behind
    if out_mask.any():
        fig.add_trace(go.Scatter(
            x=x_vals[out_mask] if isinstance(x_vals, np.ndarray) else x_vals[out_mask],
            y=y_vals[out_mask], mode="markers",
            marker=dict(size=11, symbol="x", color="lightgrey",
                        line=dict(color="#aaa", width=1.5)),
            hovertext=df["session"].to_numpy()[out_mask],
            hovertemplate="%{hovertext}<br>" + feature + " = %{y:.2f}  ⚠ outlier<extra></extra>",
            name=f"outlier (×{int(out_mask.sum())})",
        ))

    # Normal markers
    fig.add_trace(go.Scatter(
        x=x_vals[in_mask] if isinstance(x_vals, np.ndarray) else x_vals[in_mask],
        y=y_vals[in_mask], mode="markers",
        marker=dict(size=9, color=t_cal[in_mask], colorscale="Viridis", showscale=True,
                    colorbar=dict(title=cb_title)),
        hovertext=df["session"].to_numpy()[in_mask],
        hovertemplate="%{hovertext}<br>" + feature + " = %{y:.2f}<extra></extra>",
        name=feature,
    ))
    if in_mask.sum() >= 3:
        tc_in  = t_cal[in_mask]
        y_in   = y_vals[in_mask]
        z      = np.polyfit(tc_in, y_in, 1)
        yfit   = np.polyval(z, np.array([tc_in.min(), tc_in.max()]))
        x_ends = np.array([tc_in.min(), tc_in.max()])
        if x_mode != "Cycle count":
            dates_in = pd.to_datetime(df["date"].to_numpy()[in_mask])
            x_ends   = np.array([dates_in.min(), dates_in.max()])
        fig.add_trace(go.Scatter(
            x=x_ends, y=yfit, mode="lines",
            line=dict(dash="dash", color="red", width=1.5),
            name="linear trend",
        ))
    fig.update_layout(
        title=f"{feature} over {'cycles' if x_mode == 'Cycle count' else 'time'}",
        xaxis_title=x_label, yaxis_title=feature,
        margin=dict(l=50, r=50, t=60, b=40), height=390,
        legend=dict(orientation="h", y=1.12),
    )
    return fig


# ── SOH predictor analysis (bivariate + partial correlation) ───────────────────
def build_soh_predictor_table(td_soh_df: pd.DataFrame) -> pd.DataFrame:
    """
    For each testday_run feature, compute:
      - Spearman r with SOH  (bivariate: raw relationship ignoring other features)
      - Partial correlation  (unique relationship after controlling for all others)

    Partial correlation is derived from the precision matrix (Moore-Penrose inverse
    of the Spearman correlation matrix).  When the number of features exceeds the
    number of SOH observations the precision matrix is rank-deficient; p-values are
    then marked n/a but the sign and magnitude of partial correlations still indicate
    which features have a relationship with SOH that is not entirely explained by
    co-movement with the other predictors.

    Sorted by |partial corr| descending so the most uniquely predictive features
    appear first.
    """
    if td_soh_df.empty or "soh_pct" not in td_soh_df.columns:
        return pd.DataFrame()

    feat_cols = [c for c in _CORR_NUM_COLS_TREND if c in td_soh_df.columns]
    sub = td_soh_df[["soh_pct"] + feat_cols].dropna(how="any")
    if len(sub) < 4:
        return pd.DataFrame()

    n = len(sub)
    corr_mat = sub.corr(method="spearman")

    try:
        prec = np.linalg.pinv(corr_mat.values)
    except Exception:
        prec = None

    # Degrees of freedom for partial corr p-values (negative when n < p+2)
    k = len(feat_cols) - 1
    df_part = n - k - 2

    rows = []
    for j, feat in enumerate(feat_cols, start=1):
        r_biv = float(corr_mat.iloc[j, 0])
        t_biv = r_biv * np.sqrt(max(n - 2, 1) / max(1 - r_biv ** 2, 1e-15))
        p_biv = float(2 * _scipy_stats.t.sf(abs(t_biv), df=max(n - 2, 1)))

        if prec is not None:
            denom_pc = np.sqrt(max(prec[0, 0] * prec[j, j], 1e-15))
            pcorr    = float(np.clip(-prec[0, j] / denom_pc, -1.0, 1.0))
            if df_part > 0:
                t_pc  = pcorr * np.sqrt(df_part / max(1 - pcorr ** 2, 1e-15))
                p_pc  = float(2 * _scipy_stats.t.sf(abs(t_pc), df=df_part))
                p_pc_str  = f"{p_pc:.4f}"
                sig_unique = "Yes" if p_pc < 0.05 else "No"
            else:
                p_pc_str   = "n/a†"
                sig_unique = "n/a†"
        else:
            pcorr, p_pc_str, sig_unique = 0.0, "—", "—"

        rows.append({
            "Feature":                feat,
            "Spearman r (SOH)":       round(r_biv, 3),
            "p (bivariate)":          round(p_biv, 4),
            "Sig. bivariate":         "Yes" if p_biv < 0.05 else "No",
            "Partial corr (SOH)":     round(pcorr, 3),
            "p (partial)":            p_pc_str,
            "Unique predictor":       sig_unique,
        })

    if not rows:
        return pd.DataFrame()

    out = pd.DataFrame(rows)
    out["_abs"] = pd.to_numeric(out["Partial corr (SOH)"], errors="coerce").abs()
    out = out.sort_values("_abs", ascending=False).drop(columns=["_abs"]).reset_index(drop=True)
    return out


# ── ANOVA sensitivity analysis ─────────────────────────────────────────────────
def build_anova_table(trend_df: pd.DataFrame, n_groups: int = 4) -> pd.DataFrame:
    """
    One-way ANOVA sensitivity analysis on degradation trends data.

    Sessions are binned into n_groups equal-width cycle-count buckets.
    For each numerical feature we compute:
      F-statistic  – between-group variance / within-group variance
      p-value      – probability the group means are equal (small = significant trend)
      η²           – eta-squared = SS_between / SS_total (0–1, proportion of variance
                     explained by cycle group; > 0.14 = large effect)
      slope/cycle  – simple OLS slope of feature vs cycle count

    Returns rows sorted by F-statistic descending.
    """
    if trend_df.empty or "cycle" not in trend_df.columns:
        return pd.DataFrame()

    df = trend_df.copy()
    df["cycle"] = pd.to_numeric(df["cycle"], errors="coerce")
    df = df.dropna(subset=["cycle"])
    if df.empty:
        return pd.DataFrame()

    feat_cols = [c for c in _CORR_NUM_COLS_TREND if c in df.columns]
    cyc_min, cyc_max = df["cycle"].min(), df["cycle"].max()
    if cyc_max == cyc_min:
        return pd.DataFrame()

    # Equal-width bins across the full cycle range
    bins = np.linspace(cyc_min, cyc_max + 1e-9, n_groups + 1)
    df["_grp"] = pd.cut(df["cycle"], bins=bins, labels=False, include_lowest=True)

    rows = []
    for feat in feat_cols:
        sub = df[["cycle", "_grp", feat]].dropna()
        if len(sub) < 4 or sub["_grp"].nunique() < 2:
            continue

        groups = [g[feat].values for _, g in sub.groupby("_grp") if len(g) >= 2]
        if len(groups) < 2:
            continue

        f_stat, p_val = _scipy_stats.f_oneway(*groups)
        if not np.isfinite(f_stat):
            continue

        grand_mean = sub[feat].mean()
        ss_total   = float(((sub[feat] - grand_mean) ** 2).sum())
        ss_between = float(sum(len(g) * (np.mean(g) - grand_mean) ** 2 for g in groups))
        eta2       = ss_between / ss_total if ss_total > 0 else 0.0

        # OLS slope vs cycle (sign tells direction of degradation change)
        slope, intercept, r, p_slope, _ = _scipy_stats.linregress(sub["cycle"], sub[feat])

        rows.append({
            "Feature":           feat,
            "F-statistic":       round(float(f_stat), 2),
            "p-value":           round(float(p_val), 4),
            "η² effect size":    round(eta2, 3),
            "Effect":            "Large" if eta2 >= 0.14 else ("Medium" if eta2 >= 0.06 else "Small"),
            "Significant":       "Yes" if p_val < 0.05 else "No",
            "Slope / cycle":     f"{slope:+.4f}",
            "R²":                round(r ** 2, 3),
            "n sessions":        len(sub),
        })

    if not rows:
        return pd.DataFrame()
    return (
        pd.DataFrame(rows)
        .sort_values("F-statistic", ascending=False)
        .reset_index(drop=True)
    )


# ── SoC-sweep figure ───────────────────────────────────────────────────────────
def build_sweep_fig(feature: str, x_mode: str = "OCV [V]") -> go.Figure:
    """
    Plot sweep feature vs OCV or SoC, one line per sweep session.
    Groups by 'session' column (the base timestamp shared by all files in one sweep run).
    """
    global _SWEEP_DF
    fig = go.Figure()
    if _SWEEP_DF.empty or feature not in _SWEEP_DF.columns:
        return fig.update_layout(title="Click 'Compute sweep features' first")
    df = _SWEEP_DF.dropna(subset=[feature]).copy()
    if df.empty or "session" not in df.columns:
        reason = (
            "requires charge pulses ≥ 5 s — not present in sweep files"
            if feature == "CA_dVdt [mV/s]"
            else "all values are empty"
        )
        return fig.update_layout(title=f"No data for {feature}  ({reason})")

    use_soc = x_mode == "SoC [%]"
    x_col   = "soc_nominal [%]" if use_soc else "ocv_filename [V]"
    x_label = "Nominal SoC [%]"  if use_soc else "Starting OCV [V]"

    # Group by session (base timestamp = one sweep run), sorted chronologically
    sessions = sorted(df["session"].unique())
    for k, sess in enumerate(sessions):
        sub = df[df["session"] == sess].sort_values(x_col, ascending=not use_soc)
        if sub.empty:
            continue
        color     = _COLORS[k % len(_COLORS)]
        sess_date = sess[:10]  # "YYYY-MM-DD"
        # Build a human label: use block number if present, else session date
        blk_vals  = sub["block"].dropna().unique()
        label     = (f"Block {int(blk_vals[0]):02d}  ({sess_date})"
                     if len(blk_vals) > 0 else f"Sweep  {sess_date}")
        # Attach linked SOH if available
        if "linked_soh_pct" in sub.columns:
            soh_val = sub["linked_soh_pct"].dropna().iloc[0] if sub["linked_soh_pct"].notna().any() else None
            if soh_val is not None:
                label += f"  SOH={soh_val:.1f}%"

        fig.add_trace(go.Scatter(
            x=sub[x_col].to_numpy(float),
            y=sub[feature].to_numpy(float),
            mode="lines+markers", name=label,
            line=dict(color=color, width=1.8),
            marker=dict(size=10, color=color, line=dict(width=1, color="white")),
            customdata=np.stack([
                sub["date"].to_numpy(),
                sub["ocv_filename [V]"].to_numpy(),
                sub["soc_nominal [%]"].to_numpy(),
            ], axis=1),
            hovertemplate=(
                f"<b>{label}</b><br>"
                "Date: %{customdata[0]}<br>"
                "OCV = %{customdata[1]:.2f} V  (nom. SoC %{customdata[2]:.0f} %)<br>"
                f"{feature} = %{{y:.2f}}<extra></extra>"
            ),
        ))

    fig.update_layout(
        title=f"{feature} vs {x_label} — SoC sweep  ({len(sessions)} session(s))",
        xaxis_title=x_label, yaxis_title=feature,
        margin=dict(l=50, r=50, t=60, b=40), height=430,
        legend=dict(orientation="h", y=1.12),
    )
    if use_soc:
        fig.update_xaxes(range=[5, 110], tick0=10, dtick=10)
    return fig


# ── Feature correlation figures ────────────────────────────────────────────────
_CORR_NUM_COLS_TREND = [
    "DCIR_dis [mΩ]", "DCIR_chg [mΩ]", "CA_dVdt [mV/s]", "V_chg_peak [V]",
    "V_eop_dis [V]", "V_recover_dis [V]", "V_OCV [V]", "SoC_start [%]",
    "Q_dis [Ah]", "Q_chg [Ah]", "Eta_c [%]", "I_dis_peak [A]", "dur_dis_mean [s]",
    "Temp [°C]", "Humidity [%]",
]
_CORR_NUM_COLS_SWEEP = [
    "DCIR_dis [mΩ]", "V_min [V]", "V_mean_dis [V]", "V_eop_dis [V]",
    "V_recover_dis [V]", "Q_dis [Ah]", "Q_chg [Ah]",
    "E_dis [Wh]", "I_peak_dis [A]", "soc_nominal [%]", "Temp [°C]", "Humidity [%]",
]
_CORR_NUM_COLS_SOH = [
    "capacity_ah", "soh_pct",
    "dcir_1", "dcir_2", "sweep_dcir_mean",
    "temp_c", "humidity_pct",
]
# TD→SOH: testday_run features (averaged over linked runs) + SOH outcome columns
_CORR_NUM_COLS_TD_SOH = [
    "soh_pct", "capacity_ah",
    "DCIR_dis [mΩ]", "DCIR_chg [mΩ]", "CA_dVdt [mV/s]", "V_chg_peak [V]",
    "V_eop_dis [V]", "V_recover_dis [V]", "V_OCV [V]", "SoC_start [%]",
    "Q_dis [Ah]", "Q_chg [Ah]", "Eta_c [%]", "I_dis_peak [A]", "dur_dis_mean [s]",
    "Temp [°C]", "Humidity [%]",
]


def _build_td_soh_df() -> pd.DataFrame:
    """
    For each SOH measurement, average the testday_run features across the (up to 2)
    linked runs that follow it, then join with soh_pct / capacity_ah.
    Returns one row per SOH discharge.
    """
    if _SOH_DF.empty or _TREND_DF.empty:
        return pd.DataFrame()
    if "session" not in _TREND_DF.columns:
        return pd.DataFrame()

    td_keyed  = _TREND_DF.set_index("session")
    feat_cols = [c for c in _CORR_NUM_COLS_TREND if c in _TREND_DF.columns]
    rows = []
    for _, srow in _SOH_DF.iterrows():
        td_sub = []
        for key in (srow.get("run_1"), srow.get("run_2")):
            if key and key in td_keyed.index:
                td_sub.append(td_keyed.loc[key])
        if not td_sub:
            continue
        feat_vals: dict = {}
        for col in feat_cols:
            vals = [float(r[col]) for r in td_sub if col in r.index and pd.notna(r[col])]
            feat_vals[col] = float(np.mean(vals)) if vals else None
        row: dict = {
            "date":        srow["date"],
            "soh_cycle":   srow.get("soh_cycle"),
            "capacity_ah": float(srow["capacity_ah"]),
            "soh_pct":     float(srow["soh_pct"]) if pd.notna(srow.get("soh_pct")) else None,
        }
        row.update(feat_vals)
        rows.append(row)
    return pd.DataFrame(rows)


def build_corr_heatmap(df: pd.DataFrame, cols: list) -> go.Figure:
    """Spearman correlation matrix as an annotated heatmap."""
    available = [c for c in cols if c in df.columns and df[c].notna().sum() > 3]
    if len(available) < 2:
        return go.Figure().update_layout(title="Not enough data — compute features first")
    sub  = df[available].apply(pd.to_numeric, errors="coerce")
    corr = sub.corr(method="spearman").round(2)
    z    = corr.values
    labels = corr.columns.tolist()
    text   = [[f"{v:.2f}" for v in row] for row in z]
    fig = go.Figure(go.Heatmap(
        z=z, x=labels, y=labels, text=text, texttemplate="%{text}",
        colorscale="RdBu", zmid=0, zmin=-1, zmax=1,
        colorbar=dict(title="Spearman r", thickness=14),
        hovertemplate="<b>%{y}</b> vs <b>%{x}</b><br>r = %{z:.3f}<extra></extra>",
    ))
    fig.update_layout(
        title="Spearman correlation matrix",
        height=520, margin=dict(l=160, r=60, t=60, b=160),
        xaxis=dict(tickangle=-40), yaxis=dict(autorange="reversed"),
    )
    return fig


def build_scatter_pair(df: pd.DataFrame, x_col: str, y_col: str,
                       color_col: str = "date") -> go.Figure:
    """Scatter plot of two features, coloured by a third (date or cycle)."""
    if x_col not in df.columns or y_col not in df.columns:
        return go.Figure().update_layout(title=f"Column not found")
    sub = df[[x_col, y_col, color_col]].dropna() if color_col in df.columns else df[[x_col, y_col]].dropna()
    if sub.empty:
        return go.Figure().update_layout(title="No data")

    if color_col in sub.columns:
        # map colour column to numeric for continuous scale, keep label for hover
        color_vals = pd.to_numeric(sub[color_col], errors="coerce")
        if color_vals.isna().all():
            color_vals = pd.Categorical(sub[color_col]).codes.astype(float)
        marker = dict(
            color=color_vals, colorscale="Viridis", showscale=True,
            colorbar=dict(title=color_col, thickness=12),
            size=9, opacity=0.85,
        )
        customdata = sub[color_col].astype(str).values
        htmpl = (f"{x_col} = %{{x}}<br>{y_col} = %{{y}}<br>"
                 f"{color_col} = %{{customdata}}<extra></extra>")
    else:
        marker    = dict(size=9, color="#1f77b4", opacity=0.85)
        customdata = None
        htmpl     = f"{x_col} = %{{x}}<br>{y_col} = %{{y}}<extra></extra>"

    # linear trend line through all points
    x_n = pd.to_numeric(sub[x_col], errors="coerce")
    y_n = pd.to_numeric(sub[y_col], errors="coerce")
    valid = np.isfinite(x_n) & np.isfinite(y_n)
    traces = [go.Scatter(
        x=sub[x_col], y=sub[y_col],
        mode="markers", marker=marker,
        customdata=customdata, hovertemplate=htmpl, name="data",
    )]
    if valid.sum() >= 3:
        m, b = np.polyfit(x_n[valid], y_n[valid], 1)
        x_fit = np.linspace(x_n[valid].min(), x_n[valid].max(), 100)
        r = np.corrcoef(x_n[valid], y_n[valid])[0, 1]
        traces.append(go.Scatter(
            x=x_fit, y=m * x_fit + b,
            mode="lines", line=dict(color="#d62728", dash="dash", width=1.5),
            name=f"trend  r={r:.2f}", showlegend=True,
        ))

    fig = go.Figure(traces)
    fig.update_layout(
        title=f"{y_col} vs {x_col}",
        xaxis_title=x_col, yaxis_title=y_col,
        height=420, margin=dict(l=60, r=60, t=60, b=50),
        legend=dict(orientation="h", y=1.06),
    )
    return fig


# ── Export helpers ─────────────────────────────────────────────────────────────
def _style_ws(ws) -> None:
    """Navy bold header row + auto column widths."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center")
    for col_cells in ws.columns:
        width = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(width + 3, 42)


def export_excel(out_dir: str) -> str:
    """Write 3-sheet Excel summary to out_dir. Returns saved path."""
    from openpyxl import load_workbook
    os.makedirs(out_dir, exist_ok=True)
    fname    = f"battery_export_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = os.path.join(out_dir, fname)

    # ── Sheet 1: Run Summary
    src = _TREND_DF.copy() if not _TREND_DF.empty else pd.DataFrame(
        [session_summary(ts) for ts in RUNS]
    )
    src["Cycle"] = src["session"].map(CYCLE_MAP)
    _want = [
        ("session",          "Run"),
        ("date",             "Date"),
        ("Cycle",            "Cycle"),
        ("n_dis_pulses",     "# Dis pulses"),
        ("n_chg_pulses",     "# Chg pulses"),
        ("DCIR_dis [mΩ]",   "DCIR_dis [mΩ]"),
        ("DCIR_chg [mΩ]",   "DCIR_chg [mΩ]"),
        ("V_OCV [V]",        "V_OCV [V]"),
        ("SoC_start [%]",    "SoC_start [%]"),
        ("Q_dis [Ah]",       "Q_dis [Ah]"),
        ("Q_chg [Ah]",       "Q_chg [Ah]"),
        ("I_dis_peak [A]",   "I_dis_peak [A]"),
        ("dur_dis_mean [s]", "dur_dis_mean [s]"),
        ("V_eop_dis [V]",    "V_eop_dis [V]"),
        ("V_recover_dis [V]","V_recover_dis [V]"),
        ("CA_dVdt [mV/s]",   "CA_dVdt [mV/s]"),
        ("V_chg_peak [V]",   "V_chg_peak [V]"),
        ("Eta_c [%]",        "Eta_c [%]"),
    ]
    keep   = {k: v for k, v in _want if k in src.columns}
    run_df = src[list(keep)].rename(columns=keep)

    # ── Sheet 2: SOH History
    if not _SOH_DF.empty:
        s      = _SOH_DF.copy()
        cyc    = s["soh_cycle"] if "soh_cycle" in s.columns else pd.Series([None] * len(s))
        soh_df = pd.DataFrame({
            "Discharge date": s["datetime"].astype(str).str[:16],
            "Cycle":          cyc,
            "Capacity [Ah]":  s["capacity_ah"].round(3),
            "SOH [%]":        s["soh_pct"],
            "Linked run 1":   s["run_1_dt"].apply(lambda x: str(x)[:16] if x is not None else "—"),
            "DCIR_1 [mΩ]":   s["dcir_1"],
            "Linked run 2":   s["run_2_dt"].apply(lambda x: str(x)[:16] if x is not None else "—"),
            "DCIR_2 [mΩ]":   s["dcir_2"],
        })
    else:
        soh_df = pd.DataFrame(columns=[
            "Discharge date", "Cycle", "Capacity [Ah]", "SOH [%]",
            "Linked run 1", "DCIR_1 [mΩ]", "Linked run 2", "DCIR_2 [mΩ]",
        ])

    # ── Sheet 3: Cycle Map
    cycle_df = pd.DataFrame([
        {"Run": ts, "Date": ts[:10], "Time": ts[11:].replace("-", ":"), "Cycle": c}
        for ts, c in sorted(CYCLE_MAP.items(), key=lambda x: x[1])
    ])

    # ── Sheet 4: SoC Sweep
    _sweep_cols = [c for c in [
        "date", "block", "ocv_filename [V]", "soc_nominal [%]",
        "DCIR_dis [mΩ]", "V_min [V]", "V_mean_dis [V]",
        "Q_dis [Ah]", "Q_chg [Ah]", "E_dis [Wh]",
        "I_peak_dis [A]", "V_eop_dis [V]", "V_recover_dis [V]", "n_dis_pulses",
    ] if c in _SWEEP_DF.columns]
    sweep_xls = _SWEEP_DF[_sweep_cols].copy() if (not _SWEEP_DF.empty and _sweep_cols) else pd.DataFrame(
        columns=["date", "block", "ocv_filename [V]", "soc_nominal [%]", "DCIR_dis [mΩ]"]
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        run_df.to_excel(writer,    sheet_name="Run Summary", index=False)
        soh_df.to_excel(writer,    sheet_name="SOH History", index=False)
        cycle_df.to_excel(writer,  sheet_name="Cycle Map",   index=False)
        sweep_xls.to_excel(writer, sheet_name="SoC Sweep",   index=False)

    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        _style_ws(ws)
    wb.save(out_path)
    return out_path


def export_pngs(out_dir: str, figs: dict) -> tuple:
    """
    Save each go.Figure as PNG (kaleido 0.x) or interactive HTML (fallback).
    Returns (png_paths, html_paths).
    """
    import plotly.io as pio
    os.makedirs(out_dir, exist_ok=True)
    pngs, htmls = [], []
    for stem, fig in figs.items():
        if not isinstance(fig, go.Figure):
            continue
        try:
            path = os.path.join(out_dir, f"{stem}.png")
            pio.write_image(fig, path, width=2000, height=1200, scale=2)
            pngs.append(path)
        except Exception:
            path = os.path.join(out_dir, f"{stem}.html")
            fig.write_html(path, include_plotlyjs="cdn")
            htmls.append(path)
    return pngs, htmls


# ── App state ──────────────────────────────────────────────────────────────────
RUNS             = find_runs(LOG_DIR)           # post-cutoff, used for all display tabs
_ALL_RUNS        = find_all_runs(LOG_DIR)       # all runs, used only for SOH linking
FIRST_DATE       = pd.Timestamp(RUNS[0][:10]) if RUNS else pd.Timestamp.now()
CYCLE_MAP        = _compute_cycle_map()         # ts → cycle number (unique date × 10)
SOC_SWEEP_RUNS   = find_soc_sweep_runs(LOG_DIR) # SoCsweep / Block testday files


def _build_date_temp_cache(log_dir: str) -> dict:
    """
    Build {date_str: (mean_temp_C, mean_hum_pct)} from ALL sensor CSV files.
    Handles both naming conventions:
      - date AFTER  _sensor_: DegradationCycle_01_charge_sensor_YYYY-MM-DD_HH.csv
      - date BEFORE _sensor_: discharge_c5_YYYY-MM-DD_HH_sensor.csv
    """
    _any_date = re.compile(r"(\d{4}-\d{2}-\d{2})")
    temps_by_date: dict = {}
    hums_by_date:  dict = {}
    for fn in os.listdir(log_dir):
        if "sensor" not in fn.lower():
            continue
        m = _any_date.search(fn)
        if not m:
            continue
        date = m.group(1)
        try:
            # nrows=200 is enough for a stable temperature reading and avoids
            # reading multi-MB sensor files in full at startup
            df = pd.read_csv(os.path.join(log_dir, fn), nrows=200)
            df.columns = [c.strip().lower() for c in df.columns]
            if "temperature" in df.columns:
                temps_by_date.setdefault(date, []).extend(df["temperature"].dropna().tolist())
            if "humidity" in df.columns:
                hums_by_date.setdefault(date, []).extend(df["humidity"].dropna().tolist())
        except Exception:
            pass
    cache = {}
    for date in set(list(temps_by_date) + list(hums_by_date)):
        t = round(float(np.mean(temps_by_date[date])), 2) if date in temps_by_date else None
        h = round(float(np.mean(hums_by_date[date])),  1) if date in hums_by_date  else None
        cache[date] = (t, h)
    return cache


_DATE_TEMP_CACHE: dict = _build_date_temp_cache(LOG_DIR)
_TREND_DF:    pd.DataFrame = pd.DataFrame()
_SOH_DF:      pd.DataFrame = pd.DataFrame()
_SWEEP_DF:    pd.DataFrame = pd.DataFrame()
_SOH_PRED_DF: pd.DataFrame = pd.DataFrame()
_ANOVA_DF:    pd.DataFrame = pd.DataFrame()


# ── Widgets ────────────────────────────────────────────────────────────────────
run_sel = pn.widgets.Select(
    name="Test run", options=RUNS,
    value=RUNS[-1] if RUNS else None, width=340,
)
compare_sel = pn.widgets.MultiSelect(
    name="Runs to compare  (Ctrl+click for multiple)",
    options=RUNS, value=RUNS[-3:] if len(RUNS) >= 3 else RUNS,
    size=8, width=380,
)
_sweep_compare_opts = {
    f"{info['ts'][:10]}  OCV {info['ocv_filename']:.2f}V  SoC≈{info['soc_nominal']:.0f}%": info["path"]
    for info in SOC_SWEEP_RUNS
}
compare_sweep_sel = pn.widgets.MultiSelect(
    name="SoC sweep runs to compare  (Ctrl+click for multiple)",
    options=_sweep_compare_opts,
    value=[],
    size=8, width=420,
)
soh_baseline_sel = pn.widgets.Select(
    name="SOH baseline",
    options=["First valid file", "Nominal 18 Ah"],
    value="First valid file", width=200,
)
trend_feat_sel = pn.widgets.Select(
    name="Feature to trend",
    options=[
        "DCIR_dis [mΩ]", "DCIR_chg [mΩ]",
        "CA_dVdt [mV/s]", "V_chg_peak [V]",
        "V_eop_dis [V]", "V_recover_dis [V]",
        "Eta_c [%]",
        "SoC_start [%]", "V_OCV [V]",
        "Q_dis [Ah]", "Q_chg [Ah]",
        "n_dis_pulses", "n_chg_pulses",
        "I_dis_peak [A]", "dur_dis_mean [s]",
        "Temp [°C]", "Humidity [%]",
    ],
    value="DCIR_dis [mΩ]", width=280,
)

xaxis_toggle_trend = pn.widgets.RadioButtonGroup(
    name="X axis", options=["Date", "Cycle count"],
    value="Date", button_type="default", width=210,
)
anova_n_groups_sel = pn.widgets.IntSlider(
    name="Cycle groups (ANOVA bins)", start=2, end=8, step=1, value=4, width=240,
)
anova_btn   = pn.widgets.Button(name="▶  Run ANOVA sensitivity", button_type="primary", width=220)
anova_status = pn.pane.Markdown("", sizing_mode="stretch_width")
anova_tbl   = pn.widgets.Tabulator(
    pd.DataFrame(), pagination="local", page_size=20,
    height=420, sizing_mode="stretch_width",
    show_index=True,
    configuration={
        "columns": [
            {"field": "index",          "title": "#",                "width": 40},
            {"field": "Feature",        "width": 200},
            {"field": "F-statistic",    "width": 110, "hozAlign": "right"},
            {"field": "p-value",        "width": 90,  "hozAlign": "right"},
            {"field": "η² effect size", "width": 130, "hozAlign": "right"},
            {"field": "Effect",         "width": 90,  "hozAlign": "center"},
            {"field": "Significant",    "width": 100, "hozAlign": "center"},
            {"field": "Slope / cycle",  "width": 120, "hozAlign": "right"},
            {"field": "R²",             "width": 70,  "hozAlign": "right"},
            {"field": "n sessions",     "width": 95,  "hozAlign": "right"},
        ],
    },
)

# ── SOH predictor widgets ──────────────────────────────────────────────────────
soh_pred_btn    = pn.widgets.Button(
    name="▶  Run SOH predictor analysis", button_type="primary", width=250,
)
soh_pred_status = pn.pane.Markdown("", sizing_mode="stretch_width")
soh_pred_tbl    = pn.widgets.Tabulator(
    pd.DataFrame(), pagination="local", page_size=20,
    height=420, sizing_mode="stretch_width",
    show_index=True,
    configuration={
        "columns": [
            {"field": "index",                 "title": "#",   "width": 40},
            {"field": "Feature",               "width": 200},
            {"field": "Spearman r (SOH)",      "width": 145, "hozAlign": "right"},
            {"field": "p (bivariate)",         "width": 115, "hozAlign": "right"},
            {"field": "Sig. bivariate",        "width": 120, "hozAlign": "center"},
            {"field": "Partial corr (SOH)",    "width": 145, "hozAlign": "right"},
            {"field": "p (partial)",           "width": 100, "hozAlign": "right"},
            {"field": "Unique predictor",      "width": 130, "hozAlign": "center"},
        ],
    },
)


def _soh_pred_csv_callback():
    if _SOH_PRED_DF.empty:
        return io.StringIO("no data")
    buf = io.StringIO()
    _SOH_PRED_DF.to_csv(buf, index=False)
    buf.seek(0)
    return buf


def _anova_csv_callback():
    if _ANOVA_DF.empty:
        return io.StringIO("no data")
    buf = io.StringIO()
    _ANOVA_DF.to_csv(buf, index=False)
    buf.seek(0)
    return buf


soh_pred_download = pn.widgets.FileDownload(
    callback=_soh_pred_csv_callback,
    filename="soh_predictor_analysis.csv",
    button_type="default", label="Download CSV", width=160,
)
anova_download = pn.widgets.FileDownload(
    callback=_anova_csv_callback,
    filename="anova_sensitivity.csv",
    button_type="default", label="Download CSV", width=160,
)

xaxis_toggle_soh = pn.widgets.RadioButtonGroup(
    name="X axis", options=["Date", "Cycle count"],
    value="Date", button_type="default", width=210,
)

_n_sweep = len(SOC_SWEEP_RUNS)
sweep_feat_sel = pn.widgets.Select(
    name="Feature",
    options=[
        "DCIR_dis [mΩ]",
        "V_min [V]", "V_mean_dis [V]",
        "V_eop_dis [V]", "V_recover_dis [V]",
        "Q_dis [Ah]", "Q_chg [Ah]", "E_dis [Wh]",
        "I_peak_dis [A]",
        "n_dis_pulses",
    ],
    value="DCIR_dis [mΩ]", width=260,
)
xaxis_toggle_sweep = pn.widgets.RadioButtonGroup(
    name="X axis", options=["OCV [V]", "SoC [%]"],
    value="OCV [V]", button_type="default", width=180,
)
sweep_btn    = pn.widgets.Button(name="▶  Compute sweep", button_type="success", width=190)
sweep_status = pn.pane.Markdown("", sizing_mode="stretch_width")
sweep_pane   = pn.pane.Plotly(height=430, sizing_mode="stretch_width")
sweep_tbl    = pn.widgets.Tabulator(
    pd.DataFrame(), pagination="remote", page_size=15,
    height=320, sizing_mode="stretch_width",
)

_trend_corr_cols  = _CORR_NUM_COLS_TREND
_sweep_corr_cols  = _CORR_NUM_COLS_SWEEP
corr_dataset_sel  = pn.widgets.RadioButtonGroup(
    name="Dataset",
    options=["Degradation Trends", "SOH History", "TD → SOH (linked)", "SoC Sweep"],
    value="Degradation Trends", button_type="default", width=560,
)
corr_x_sel = pn.widgets.Select(
    name="Scatter X", options=_trend_corr_cols, value="DCIR_dis [mΩ]", width=240,
)
corr_y_sel = pn.widgets.Select(
    name="Scatter Y", options=_trend_corr_cols, value="CA_dVdt [mV/s]", width=240,
)
corr_color_sel = pn.widgets.Select(
    name="Colour by", options=["date", "cycle"], value="cycle", width=160,
)
corr_btn   = pn.widgets.Button(name="▶  Compute correlations", button_type="success", width=210)
corr_heatmap_pane = pn.pane.Plotly(height=540, sizing_mode="stretch_width")
corr_scatter_pane = pn.pane.Plotly(height=440, sizing_mode="stretch_width")

export_dir_input = pn.widgets.TextInput(
    name="Output folder",
    value=os.path.join(BASE_DIR, "exports"),
    width=520,
)
do_pngs_toggle = pn.widgets.Checkbox(
    name="Also export PNG images of current plots  (requires kaleido — pip install kaleido)",
    value=True,
)
export_btn       = pn.widgets.Button(name="▼  Export",          button_type="primary", width=140)
export_status_md = pn.pane.Markdown("", sizing_mode="stretch_width")

load_btn    = pn.widgets.Button(name="↻  Load",           button_type="primary", width=120)
compare_btn = pn.widgets.Button(name="⧉  Compare",        button_type="warning",  width=130)
soh_btn     = pn.widgets.Button(name="▶  Compute SOH",    button_type="success",  width=160)
trend_btn   = pn.widgets.Button(name="▶  Compute trends", button_type="success",  width=170)

status      = pn.pane.Markdown("", sizing_mode="stretch_width")
soh_status  = pn.pane.Markdown("", sizing_mode="stretch_width")

sig_pane     = pn.pane.Plotly(height=430, sizing_mode="stretch_width")
compare_pane = pn.pane.Plotly(height=550, sizing_mode="stretch_width")
soh_pane     = pn.pane.Plotly(height=580, sizing_mode="stretch_width")
trend_pane   = pn.pane.Plotly(height=390, sizing_mode="stretch_width")

pulse_tbl = pn.widgets.Tabulator(
    pd.DataFrame(), pagination="remote", page_size=15,
    height=360, sizing_mode="stretch_width",
)
soh_tbl = pn.widgets.Tabulator(
    pd.DataFrame(), pagination="remote", page_size=20,
    height=320, sizing_mode="stretch_width",
)


# ── Callbacks ──────────────────────────────────────────────────────────────────
def _load_run(event=None):
    ts = run_sel.value
    if not ts:
        return
    status.object   = f"Loading **{ts}** …"
    sig_pane.object = build_signal_fig(ts)
    sig_pane.config = _plotly_cfg(f"testday_run_{ts}")
    pulse_tbl.value = build_pulse_table(ts)
    status.object   = f"Loaded **{ts}** — {len(pulse_tbl.value)} pulse event(s)"


def _compare_runs(event=None):
    selected       = compare_sel.value or []
    sweep_selected = compare_sweep_sel.value or []
    if not selected and not sweep_selected:
        status.object = "Select at least one run."
        return
    n_total = len(selected) + len(sweep_selected)
    status.object        = f"Plotting {n_total} trace(s) …"
    compare_pane.object  = build_compare_fig(selected, sweep_selected)
    stems                = "_vs_".join(ts.replace("-", "").replace("_", "") for ts in selected[:3])
    compare_pane.config  = _plotly_cfg(f"compare_{stems}")
    status.object        = f"Comparison ready — {n_total} trace(s)"


def _compute_soh(event=None):
    global _SOH_DF, _SWEEP_DF
    baseline = "first" if soh_baseline_sel.value == "First valid file" else "nominal"
    soh_status.object = "Computing SOH and linking to testday_run and SoCsweep files …"
    soh_df = compute_soh_series(nominal_ah=18.0, baseline=baseline)
    if soh_df.empty:
        soh_status.object = "No valid discharge_c5 files found."
        return
    soh_df = link_soh_to_runs(soh_df)
    soh_df = link_soh_to_sweeps(soh_df)          # adds sweep_session, sweep_dcir_mean, …
    # If sweep features exist, annotate sweep rows with the downstream SOH value
    if not _SWEEP_DF.empty:
        _SWEEP_DF = _annotate_sweeps_with_soh(_SWEEP_DF, soh_df)
        if not _SWEEP_DF.empty:
            sweep_pane.object = build_sweep_fig(sweep_feat_sel.value, xaxis_toggle_sweep.value)
    _SOH_DF = soh_df
    soh_pane.object = build_soh_fig(soh_df, xaxis_toggle_soh.value)
    soh_pane.config = _plotly_cfg("battery_soh_history")
    soh_tbl.value   = build_soh_table(soh_df)
    n_linked  = int(soh_df["run_1"].notna().sum())
    n_sweeps  = int(soh_df["sweep_session"].notna().sum()) if "sweep_session" in soh_df.columns else 0
    soh_status.object = (
        f"SOH ready — {len(soh_df)} discharge files · "
        f"{n_linked} linked to testday_runs · "
        f"{n_sweeps} linked to SoCsweep sessions"
    )


def _compute_trends(event=None):
    global _TREND_DF
    status.object = "Computing features for all runs … (~10 s)"
    _TREND_DF = pd.DataFrame([session_summary(ts) for ts in RUNS])
    _TREND_DF["cycle"] = _TREND_DF["session"].map(CYCLE_MAP)
    trend_pane.object = build_trend_fig(trend_feat_sel.value, xaxis_toggle_trend.value)
    trend_pane.config = _plotly_cfg("battery_degradation_trends")
    status.object     = f"Trends ready — {len(_TREND_DF)} runs"


def _update_trend_feature(event=None):
    trend_pane.object = build_trend_fig(trend_feat_sel.value, xaxis_toggle_trend.value)


def _update_trend_xaxis(event=None):
    trend_pane.object = build_trend_fig(trend_feat_sel.value, xaxis_toggle_trend.value)


def _run_soh_predictor(event=None):
    global _SOH_PRED_DF
    if _TREND_DF.empty:
        soh_pred_status.object = "**Compute trends first** (▶ Compute trends above)."
        return
    if _SOH_DF.empty:
        soh_pred_status.object = "**Compute SOH first** (▶ Compute SOH in the SOH History tab)."
        return
    soh_pred_status.object = "_Building TD→SOH dataset and running analysis …_"
    td_soh = _build_td_soh_df()
    if td_soh.empty:
        soh_pred_status.object = "No linked TD→SOH rows found."
        return
    result = build_soh_predictor_table(td_soh)
    if result.empty:
        soh_pred_status.object = "Not enough data."
        return
    n_obs = len(td_soh.dropna(how="any", subset=["soh_pct"] + [c for c in _CORR_NUM_COLS_TREND if c in td_soh.columns]))
    n_feat = len([c for c in _CORR_NUM_COLS_TREND if c in td_soh.columns])
    df_part = n_obs - n_feat - 2
    note = (
        f"  †  Partial p-values unavailable (df = {df_part} < 0): "
        f"{n_obs} SOH observations, {n_feat} features controlled.  "
        "Partial corr rankings are still informative."
        if df_part <= 0 else ""
    )
    n_sig_biv = int((result["Sig. bivariate"] == "Yes").sum())
    soh_pred_status.object = (
        f"**{n_obs} SOH observations** · **{n_feat} features** · "
        f"**{n_sig_biv} significant bivariate** (p < 0.05)" + note
    )
    _SOH_PRED_DF = result
    soh_pred_tbl.value = result


def _run_anova(event=None):
    global _ANOVA_DF
    if _TREND_DF.empty:
        anova_status.object = "**Compute trends first** (click ▶ Compute trends above)."
        return
    anova_status.object = "_Running ANOVA …_"
    result = build_anova_table(_TREND_DF, n_groups=anova_n_groups_sel.value)
    if result.empty:
        anova_status.object = "No results — not enough sessions per group."
        return
    n_sig = int((result["Significant"] == "Yes").sum())
    n_large = int((result["Effect"] == "Large").sum())
    anova_status.object = (
        f"**{len(result)} features analysed** · "
        f"**{n_sig} significant** (p < 0.05) · "
        f"**{n_large} large effect** (η² ≥ 0.14) · "
        f"Cycle range split into **{anova_n_groups_sel.value} groups**"
    )
    _ANOVA_DF = result
    anova_tbl.value = result


def _update_soh_xaxis(event=None):
    if not _SOH_DF.empty:
        soh_pane.object = build_soh_fig(_SOH_DF, xaxis_toggle_soh.value)


def _compute_sweep(event=None):
    global _SWEEP_DF, _SOH_DF
    if not SOC_SWEEP_RUNS:
        sweep_status.object = "No SoC sweep files found in the log folder."
        return
    n_sessions = len({info["session"] for info in SOC_SWEEP_RUNS})
    sweep_status.object = (
        f"Computing features for {len(SOC_SWEEP_RUNS)} sweep run(s) "
        f"across {n_sessions} session(s) …"
    )
    _SWEEP_DF = pd.DataFrame([sweep_session_summary(info) for info in SOC_SWEEP_RUNS])

    # If SOH is already computed, cross-link sweep sessions ↔ SOH values
    if not _SOH_DF.empty:
        _SOH_DF = link_soh_to_sweeps(_SOH_DF)
        _SWEEP_DF = _annotate_sweeps_with_soh(_SWEEP_DF, _SOH_DF)
        soh_pane.object = build_soh_fig(_SOH_DF, xaxis_toggle_soh.value)
        soh_tbl.value   = build_soh_table(_SOH_DF)

    sweep_pane.object = build_sweep_fig(sweep_feat_sel.value, xaxis_toggle_sweep.value)
    sweep_pane.config  = _plotly_cfg("battery_soc_sweep")
    tbl_cols = [c for c in [
        "session", "date", "block", "ocv_filename [V]", "soc_nominal [%]",
        "linked_soh_pct",
        "DCIR_dis [mΩ]", "V_min [V]", "V_mean_dis [V]",
        "Q_dis [Ah]", "E_dis [Wh]", "n_dis_pulses",
    ] if c in _SWEEP_DF.columns]
    sweep_tbl.value = _SWEEP_DF[tbl_cols].copy()
    sweep_status.object = (
        f"Sweep ready — {len(_SWEEP_DF)} run(s)  ·  {n_sessions} session(s)"
    )


def _update_sweep_feature(event=None):
    if not _SWEEP_DF.empty:
        sweep_pane.object = build_sweep_fig(sweep_feat_sel.value, xaxis_toggle_sweep.value)


def _update_sweep_xaxis(event=None):
    if not _SWEEP_DF.empty:
        sweep_pane.object = build_sweep_fig(sweep_feat_sel.value, xaxis_toggle_sweep.value)


def _get_corr_df() -> tuple:
    """Return (df, cols, color_opts) for the currently selected dataset."""
    sel = corr_dataset_sel.value
    if sel == "SoC Sweep":
        df         = _SWEEP_DF.copy() if not _SWEEP_DF.empty else pd.DataFrame()
        cols       = _CORR_NUM_COLS_SWEEP
        color_opts = ["date", "soc_nominal [%]", "session"]
    elif sel == "SOH History":
        df         = _SOH_DF.copy() if not _SOH_DF.empty else pd.DataFrame()
        cols       = _CORR_NUM_COLS_SOH
        color_opts = ["date", "soh_cycle"]
    elif sel == "TD → SOH (linked)":
        df         = _build_td_soh_df()
        cols       = _CORR_NUM_COLS_TD_SOH
        color_opts = ["date", "soh_cycle"]
    else:  # Degradation Trends
        df         = _TREND_DF.copy() if not _TREND_DF.empty else pd.DataFrame()
        cols       = _CORR_NUM_COLS_TREND
        color_opts = ["date", "cycle"]
    return df, cols, color_opts


def _refresh_corr(event=None):
    df, cols, color_opts = _get_corr_df()
    available = [c for c in cols if not df.empty and c in df.columns]
    # Compute new values before touching options to avoid cascade None states
    new_x = corr_x_sel.value if corr_x_sel.value in available else (available[0] if available else None)
    new_y = corr_y_sel.value if corr_y_sel.value in available else (available[1] if len(available) > 1 else new_x)
    new_col = corr_color_sel.value if corr_color_sel.value in color_opts else (color_opts[0] if color_opts else None)
    # Update options and value together to avoid watcher firing with None value
    corr_x_sel.param.update(options=available, value=new_x)
    corr_y_sel.param.update(options=available, value=new_y)
    corr_color_sel.param.update(options=color_opts, value=new_col)
    corr_heatmap_pane.object = build_corr_heatmap(df, cols)
    corr_heatmap_pane.config = _plotly_cfg("feature_correlations")
    _refresh_corr_scatter()


def _refresh_corr_scatter(event=None):
    df, cols, _ = _get_corr_df()
    if df.empty:
        corr_scatter_pane.object = go.Figure().update_layout(title="Compute features first")
        return
    x   = corr_x_sel.value
    y   = corr_y_sel.value
    if not x or not y:
        return
    col = corr_color_sel.value or "date"
    color_col = col if col in df.columns else "date"
    corr_scatter_pane.object = build_scatter_pair(df, x, y, color_col)
    corr_scatter_pane.config = _plotly_cfg(f"scatter_{x[:8]}_vs_{y[:8]}")


def _run_export(event=None):
    out_dir = export_dir_input.value.strip() or os.path.join(BASE_DIR, "exports")
    lines   = []
    try:
        export_status_md.object = "_Exporting Excel …_"
        xls = export_excel(out_dir)
        lines.append(f"**Excel saved:** `{xls}`")
    except Exception as e:
        export_status_md.object = f"**Excel export failed:** {e}"
        return

    if do_pngs_toggle.value:
        feat_clean = trend_feat_sel.value.split("[")[0].strip().replace(" ", "_")
        figs = {
            f"01_signal_{run_sel.value}":  sig_pane.object,
            "02_compare_runs":             compare_pane.object,
            "03_soh_history":              soh_pane.object,
            f"04_trend_{feat_clean}":      trend_pane.object,
            "05_soc_sweep":                sweep_pane.object,
        }
        try:
            pngs, htmls = export_pngs(out_dir, figs)
            parts = []
            if pngs:
                parts.append(f"**{len(pngs)} PNG(s) saved**")
            if htmls:
                parts.append(
                    f"**{len(htmls)} HTML(s) saved** *(PNG needs kaleido 0.x — "
                    f"run `pip install \"kaleido==0.2.1\"` once to get PNGs next time)*"
                )
            lines.append(" · ".join(parts) + f"\n→ `{out_dir}`")
        except Exception as e:
            lines.append(f"**Image export failed:** {e}")

    export_status_md.object = "\n\n".join(lines)


run_sel.param.watch(_load_run, "value")
load_btn.on_click(_load_run)
compare_btn.on_click(_compare_runs)
soh_btn.on_click(_compute_soh)
trend_btn.on_click(_compute_trends)
trend_feat_sel.param.watch(_update_trend_feature, "value")
xaxis_toggle_trend.param.watch(_update_trend_xaxis, "value")
anova_btn.on_click(_run_anova)
soh_pred_btn.on_click(_run_soh_predictor)
xaxis_toggle_soh.param.watch(_update_soh_xaxis, "value")
sweep_btn.on_click(_compute_sweep)
sweep_feat_sel.param.watch(_update_sweep_feature, "value")
xaxis_toggle_sweep.param.watch(_update_sweep_xaxis, "value")
corr_btn.on_click(_refresh_corr)
corr_dataset_sel.param.watch(_refresh_corr, "value")
corr_x_sel.param.watch(_refresh_corr_scatter, "value")
corr_y_sel.param.watch(_refresh_corr_scatter, "value")
corr_color_sel.param.watch(_refresh_corr_scatter, "value")
export_btn.on_click(_run_export)

_load_run()


# ── Feature legend ─────────────────────────────────────────────────────────────
_LEGEND = pn.pane.Markdown(r"""
**Feature legend**

| Symbol | Meaning |
|--------|---------|
| **DCIR_dis** | DC internal resistance from discharge pulse: ΔV / \|I\| × 1000 [mΩ]. Computed over first 3 s of pulse. |
| **DCIR_chg** | DC internal resistance from charge pulse [mΩ]. |
| **CA_dVdt** | Charge acceptance: voltage rise rate during CC charge pulse [mV/s]. *Lower = better acceptance = healthier battery.* |
| **V_chg_peak** | Mean peak voltage reached during charge pulses [V]. Rises as battery degrades. |
| **V_eop_dis** | End-of-pulse voltage: mean voltage in the last 2 s of each discharge pulse [V]. Reflects voltage sag depth. |
| **V_recover_dis** | Recovery voltage: mean voltage in the 2 s rest window after each discharge pulse [V]. Shows kinetic recovery. |
| **Eta_c** | Coulombic efficiency: Q_chg / Q_dis × 100 [%]. Ratio of charge returned to charge drawn per test session. |
| **V_OCV** | Open-circuit voltage during rest before each pulse [V] |
| **SoC** | State of Charge estimated from V_OCV via the VRLA/AGM resting-voltage table [%] *(approximate — battery must be rested ≥ 1 h for accuracy)* |
| **ΔV** | Voltage swing during the pulse = \|V_pulse − V_OCV\| [V] |
| **Q_dis / Q_chg** | Total charge passed during all discharge / charge pulses in the session [Ah] |
| **I_peak** | Peak current of the discharge pulse [A] — negative = discharge |
| **dur_dis** | Average duration of discharge pulses [s] |
""", sizing_mode="stretch_width")


# ── Test-Day v2 (Beta) tab ──────────────────────────────────────────────────────
# New "realistic test day" profile format: replays a scripted event sequence
# (ocv_window, wakeup_load, glow_plug_like_load, crank_pulse x2, recovery_rest,
# alternator_charge, ramp_like_load, driving_aux_load) and logs which event was
# active per row via Event_Type/Event_Index columns, instead of the old
# threshold-detected rest/discharge/charge pulses used elsewhere in this
# dashboard. See testday_v2_features.py for the extraction logic.
#
# No file on disk has this schema yet (backend.py hasn't rolled it out) — this
# tab is built ahead of that so it starts populating automatically once it does.
_V2_DF = pd.DataFrame(columns=[
    "battery_id", "block", "ocv_label_V", "kind", "timestamp", "filename",
])

v2_battery_sel = pn.widgets.Select(name="Battery ID", options=["All"], value="All", width=200)
v2_scan_btn = pn.widgets.Button(name="▶  Scan for v2 profile runs", button_type="primary", width=220)
v2_status = pn.pane.Markdown(
    "Not scanned yet — click **Scan for v2 profile runs**.", sizing_mode="stretch_width",
)
v2_tbl = pn.widgets.Tabulator(
    pd.DataFrame(), pagination="local", page_size=15,
    height=360, sizing_mode="stretch_width",
)
v2_crank_pane = pn.pane.Plotly(go.Figure(), config={"responsive": True}, sizing_mode="stretch_width")
v2_soh_pane = pn.pane.Plotly(go.Figure(), config={"responsive": True}, sizing_mode="stretch_width")


def _build_v2_crank_fig(df: pd.DataFrame) -> go.Figure:
    fig = make_subplots(rows=1, cols=1, subplot_titles=["Crank apparent R_int vs run"])
    for label, colour in [("cold", "#1f77b4"), ("hot", "#EF553B")]:
        col = f"crank_{label}_R_int_apparent_mohm"
        if col not in df.columns:
            continue
        sub = df.dropna(subset=[col])
        if sub.empty:
            continue
        fig.add_trace(go.Scatter(
            x=sub["timestamp"], y=sub[col], mode="markers+lines",
            name=f"{label} crank", marker=dict(size=8, color=colour),
            text=sub["battery_id"],
        ))
    fig.update_yaxes(title_text="R_int_apparent [mΩ]")
    fig.update_xaxes(title_text="Run timestamp")
    fig.update_layout(height=380, template="plotly_white", title="Crank R_int trend (by battery_id)")
    return fig


def _build_v2_block_soh_fig(df: pd.DataFrame) -> go.Figure:
    fig = go.Figure()
    if "block_c5_capacity_ah" in df.columns:
        sub = df.dropna(subset=["block_c5_capacity_ah", "block"]).drop_duplicates(
            subset=["battery_id", "block"]
        )
        for bid, g in sub.groupby("battery_id"):
            g = g.sort_values("block")
            fig.add_trace(go.Scatter(
                x=g["block"], y=g["block_c5_capacity_ah"], mode="markers+lines",
                name=bid,
            ))
    fig.update_xaxes(title_text="Block #")
    fig.update_yaxes(title_text="C/5 capacity [Ah]")
    fig.update_layout(height=380, template="plotly_white", title="Block-level measured C/5 capacity (SOH ground truth)")
    return fig


def _scan_v2(event=None):
    global _V2_DF
    full_df = v2feat.build_dataset(LOG_DIR)
    ids = sorted(full_df.get("battery_id", pd.Series(dtype=str)).dropna().unique().tolist())
    v2_battery_sel.options = ["All"] + ids

    bid = None if v2_battery_sel.value in (None, "All") else v2_battery_sel.value
    _V2_DF = full_df if bid is None else full_df[full_df["battery_id"] == bid].reset_index(drop=True)
    _V2_DF.attrs = full_df.attrs

    n_legacy = _V2_DF.attrs.get("n_skipped_legacy_schema", 0)
    n_corrupt = _V2_DF.attrs.get("n_skipped_corrupt_hash", 0)

    if _V2_DF.empty:
        v2_status.object = (
            f"**No v2-format test-day profile runs found** "
            f"({n_legacy} legacy-schema file(s) skipped"
            + (f", {n_corrupt} corrupt Profile_Hash file(s) skipped" if n_corrupt else "")
            + "). This tab will populate automatically once backend.py starts logging "
            "the `Event_Type` / `Event_Index` / `Profile_Hash` / `Battery_ID` columns."
        )
    else:
        n_batt = _V2_DF["battery_id"].nunique()
        v2_status.object = (
            f"**{len(_V2_DF)} v2 run(s)** found across **{n_batt} battery ID(s)**  ·  "
            f"{n_legacy} legacy-schema file(s) skipped"
            + (f", {n_corrupt} corrupt Profile_Hash file(s) skipped" if n_corrupt else "")
            + ". Remember: never pool rows across different battery_id values."
        )

    v2_tbl.value = _V2_DF
    v2_crank_pane.object = _build_v2_crank_fig(_V2_DF)
    v2_soh_pane.object = _build_v2_block_soh_fig(_V2_DF)


v2_scan_btn.on_click(_scan_v2)

v2_tab = pn.Column(
    pn.pane.Markdown("## Test-Day v2 (Beta)"),
    pn.pane.Markdown(
        "Analyses the new event-scripted test-day profile "
        "(`*_testday_bdps_*.csv` / `*_profile_single_*_bdps.csv` with "
        "`Event_Type`/`Event_Index`/`Profile_Hash`/`Battery_ID` columns), which replaces the "
        "old current-threshold pulse detection with an explicit event script: "
        "`ocv_window → wakeup_load → glow_plug_like_load → crank_pulse (cold) → recovery_rest → "
        "glow_plug_like_load → crank_pulse (hot) → recovery_rest → alternator_charge → "
        "ramp_like_load / driving_aux_load`.\n\n"
        "Two independent labels come out of this data — don't conflate them: "
        "**per-run** starting OCV (parsed from the filename) for a \"what was OCV at run start\" model, "
        "and **per-block** measured C/5 Ah capacity (from `Block_<nn>_SOH_C5_bdps_*.csv`) as the real SOH "
        "ground truth, shared by all runs in that block.\n\n"
        "*Beta:* no logged file currently has this schema — once it does, click **Scan** to pick it up."
    ),
    pn.Row(v2_battery_sel, v2_scan_btn, sizing_mode="stretch_width"),
    v2_status,
    v2_crank_pane,
    v2_soh_pane,
    pn.pane.Markdown("### Per-run feature table"),
    v2_tbl,
    sizing_mode="stretch_width",
)


# ── Layout ─────────────────────────────────────────────────────────────────────
inspector_tab = pn.Column(
    pn.pane.Markdown("## Test Run Inspector"),
    pn.pane.Markdown(
        "Select a `testday_run` file (≥ 2026-01-27). Detected pulses: "
        "**orange** = discharge, **blue** = charge. "
        "Red/green numbers = DCIR [mΩ]. Title shows estimated SoC from OCV."
    ),
    pn.Row(run_sel, load_btn, sizing_mode="stretch_width"),
    status,
    sig_pane,
    pn.pane.Markdown("### Pulse event table"),
    pulse_tbl,
    _LEGEND,
    sizing_mode="stretch_width",
)

compare_tab = pn.Column(
    pn.pane.Markdown("## Compare Runs"),
    pn.pane.Markdown(
        "Ctrl+click (or Shift+click) to select multiple runs, then click **Compare**. "
        "All traces share a common elapsed-time axis. "
        "Starting SoC is shown in each run's legend entry."
    ),
    pn.Row(
        pn.Column(compare_sel),
        pn.Column(compare_sweep_sel),
        compare_btn,
        sizing_mode="stretch_width",
    ),
    compare_pane,
    sizing_mode="stretch_width",
)

soh_tab = pn.Column(
    pn.pane.Markdown("## SOH History"),
    pn.pane.Markdown(
        "SOH is computed from `discharge_c5_*_bdps.csv` files (≥ 20 kB). "
        "**Each SOH measurement is linked to the 2 nearest testday_run files that follow it.** "
        "The lower panel shows the DCIR from those linked runs at their actual dates — "
        "compare how DCIR (real-world impedance) tracks the formal SOH."
    ),
    pn.Row(soh_baseline_sel,
           pn.Column(pn.pane.Markdown("**X axis**", margin=(0, 0, 2, 0)), xaxis_toggle_soh),
           soh_btn,
           sizing_mode="stretch_width"),
    soh_status,
    soh_pane,
    pn.pane.Markdown("### SOH linkage table"),
    soh_tbl,
    sizing_mode="stretch_width",
)

trends_tab = pn.Column(
    pn.pane.Markdown("## Degradation Trends"),
    pn.pane.Markdown(
        "Computes a one-row summary for every testday_run file and plots the "
        "selected feature over calendar date or cycle count. Dashed red line = linear trend."
    ),
    pn.Row(trend_feat_sel,
           pn.Column(pn.pane.Markdown("**X axis**", margin=(0, 0, 2, 0)), xaxis_toggle_trend),
           trend_btn,
           sizing_mode="stretch_width"),
    trend_pane,
    pn.layout.Divider(),
    pn.pane.Markdown("### SOH Predictor Analysis  *(which features independently predict SOH?)*"),
    pn.pane.Markdown(
        "For each testday_run feature, two correlations with SOH are shown:\n\n"
        "- **Spearman r (bivariate)** — raw correlation with SOH ignoring all other features. "
        "A high value means the feature tracks SOH, but it may simply be because "
        "both change together as the battery ages.\n"
        "- **Partial corr (SOH)** — correlation with SOH *after removing the effect of all other features*. "
        "A high partial correlation means this feature adds unique predictive information about SOH "
        "beyond what the other features already explain. "
        "Features with high bivariate r but near-zero partial corr are redundant with other predictors.\n\n"
        "*Requires: **▶ Compute trends** + **▶ Compute SOH** (SOH History tab) both run first.*  \n"
        "†  Partial p-values require more observations than features; "
        "with small n the ranking of partial correlations is still informative."
    ),
    pn.Row(soh_pred_btn, soh_pred_download, sizing_mode="stretch_width"),
    soh_pred_status,
    soh_pred_tbl,
    pn.layout.Divider(),
    pn.pane.Markdown("### ANOVA Sensitivity Analysis  *(which features change significantly over the degradation cycle?)*"),
    pn.pane.Markdown(
        "One-way ANOVA tests whether each feature's mean differs significantly across "
        "equal-width cycle-count bins. This does **not** link to SOH — it only measures "
        "how much each feature changes over time.\n\n"
        "- **F-statistic** — higher = more between-group separation vs within-group noise.\n"
        "- **p-value** — below 0.05 means the trend is statistically significant.\n"
        "- **η² (eta-squared)** — effect size: proportion of total variance explained by cycle group "
        "(Small < 0.06 · Medium 0.06–0.14 · Large ≥ 0.14).\n"
        "- **Slope / cycle** — OLS trend direction and magnitude per cycle step.\n\n"
        "*Requires: **▶ Compute trends** first.*"
    ),
    pn.Row(anova_n_groups_sel, anova_btn, anova_download, sizing_mode="stretch_width"),
    anova_status,
    anova_tbl,
    sizing_mode="stretch_width",
)

sweep_tab = pn.Column(
    pn.pane.Markdown("## SoC Sweep Analysis"),
    pn.pane.Markdown(
        f"Analyses the test-day profile run at **{_n_sweep} different SoC levels** per degradation block.\n\n"
        "Each run starts at a different OCV (set by the step-down sequence on the Pi). "
        "Plotting a feature vs OCV shows how battery behaviour changes with SoC — and "
        "stacking multiple blocks reveals how that SoC-dependence evolves as the battery degrades."
    ),
    pn.Row(
        sweep_feat_sel,
        pn.Column(pn.pane.Markdown("**X axis**", margin=(0, 0, 2, 0)), xaxis_toggle_sweep),
        sweep_btn,
        sizing_mode="stretch_width",
    ),
    sweep_status,
    sweep_pane,
    pn.pane.Markdown("### Sweep results table"),
    sweep_tbl,
    sizing_mode="stretch_width",
)

export_tab = pn.Column(
    pn.pane.Markdown("## Export to Excel & PNG"),
    pn.pane.Markdown(
        "Creates a timestamped Excel file and optionally saves the currently "
        "displayed plots as high-resolution PNG files.\n\n"
        "| Sheet | Contents |\n"
        "|---|---|\n"
        "| **Run Summary** | Feature values per testday_run + cycle number |\n"
        "| **SOH History** | Capacity, SOH%, linked runs and DCIR |\n"
        "| **Cycle Map** | Run timestamps → cycle count |\n"
        "| **SoC Sweep** | Features per sweep run (OCV, SoC, DCIR, V_min, …) |\n\n"
        "> Tip — click **Compute trends**, **Compute SOH**, and **Compute sweep** first so the Excel "
        "contains fully computed data. If those are skipped, Run Summary is computed "
        "on the fly (may take ~10 s)."
    ),
    pn.Row(export_dir_input, sizing_mode="stretch_width"),
    do_pngs_toggle,
    pn.Row(export_btn),
    export_status_md,
    sizing_mode="stretch_width",
)

corr_tab = pn.Column(
    pn.pane.Markdown("## Feature Correlations"),
    pn.pane.Markdown(
        "Spearman correlation between all computed features — values near **+1 / −1** mean "
        "strong co-movement; near **0** means independent.\n\n"
        "- **Degradation Trends** — one row per testday_run session (pulse features + temp/humidity)\n"
        "- **SOH History** — one row per C/5 discharge (capacity, SOH%, linked DCIR, temp) — *requires Compute SOH*\n"
        "- **TD → SOH (linked)** — testday_run features averaged from the 2 runs after each SOH, "
        "correlated against `soh_pct` / `capacity_ah` — *requires both Compute Trends and Compute SOH*\n"
        "- **SoC Sweep** — one row per SoCsweep run\n\n"
        "Click a cell in the heatmap, then explore the pair in the scatter plot below."
    ),
    pn.Row(
        pn.Column(pn.pane.Markdown("**Dataset**", margin=(0,0,2,0)), corr_dataset_sel),
        corr_btn,
        sizing_mode="stretch_width",
    ),
    corr_heatmap_pane,
    pn.pane.Markdown("### Pairwise scatter"),
    pn.Row(corr_x_sel, corr_y_sel, corr_color_sel, sizing_mode="stretch_width"),
    corr_scatter_pane,
    sizing_mode="stretch_width",
)

# ── File Inventory tab ─────────────────────────────────────────────────────────
_DEGRAD_SENSOR_PAT = re.compile(
    r"^DegradationCycle_(\d+)_(charge|discharge)_sensor_"
    r"(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2}-\d{2})\.csv$"
)
_DEGR_BDPS_PAT = re.compile(
    r"^Degr_(\d+)_(charge|discharge)_bdps_"
    r"(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2}-\d{2})\.csv$"
)
_DEGR_SENSOR_PAT = re.compile(
    r"^Degr_(\d+)_(charge|discharge)_sensor_"
    r"(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2}-\d{2})\.csv$"
)


def build_file_inventory(log_dir: str) -> pd.DataFrame:
    """
    One row per primary data file (or one summary row per DegradationCycle session).
    Filters to files actually used in the dashboard; excludes full_charge, cycle_,
    and small discharge_c5 fragments.
    """
    rows = []

    # Pre-build set of Degr_ sensor keys for quick lookup
    degr_sensor_keys: set = set()
    for fn in os.listdir(log_dir):
        m = _DEGR_SENSOR_PAT.match(fn)
        if m:
            degr_sensor_keys.add((m.group(1), m.group(2), m.group(3), m.group(4)))

    _INV_TESTDAY_CUTOFF = pd.Timestamp("2025-12-09")  # earlier runs were test/debug sessions

    # 1 — testday_run bdps files (from 2025-12-09 onwards; earlier were test sessions)
    for fn in sorted(os.listdir(log_dir)):
        m = _TESTDAY_PAT.match(fn)
        if not m:
            continue
        ts   = m.group(1)
        date = ts[:10]
        if pd.Timestamp(date) < _INV_TESTDAY_CUTOFF:
            continue
        time = ts[11:]
        path = os.path.join(log_dir, fn)
        size_kb = os.path.getsize(path) // 1024
        temp_c, hum = _DATE_TEMP_CACHE.get(date, (None, None))
        post  = pd.Timestamp(date) >= TESTDAY_CUTOFF
        rows.append({
            "Date":         date,
            "Time":         time.replace("-", ":"),
            "Filename":     fn,
            "Test type":    "Pulse test",
            "Used in":      "Run Inspector / Compare / Trends" if post else "SOH linking only",
            "Size (KB)":    size_kb,
            "Sensor":       "—",
            "Temp [°C]":    round(temp_c, 1) if temp_c is not None else "—",
            "Humidity [%]": round(float(hum), 0) if hum is not None else "—",
        })

    # 2 — discharge_c5 bdps files (valid only, ≥ SOH_MIN_BYTES)
    for fn in sorted(os.listdir(log_dir)):
        m = _DIS_C5_PAT.match(fn)
        if not m:
            continue
        path = os.path.join(log_dir, fn)
        size = os.path.getsize(path)
        if size < SOH_MIN_BYTES:
            continue
        date, time = m.group(1), m.group(2)
        size_kb = size // 1024
        temp_c, hum = _DATE_TEMP_CACHE.get(date, (None, None))
        sensor_fn = f"discharge_c5_{date}_{time}_sensor.csv"
        has_sensor = os.path.exists(os.path.join(log_dir, sensor_fn))
        rows.append({
            "Date":         date,
            "Time":         time.replace("-", ":"),
            "Filename":     fn,
            "Test type":    "SOH test (C/5 discharge)",
            "Used in":      "SOH History",
            "Size (KB)":    size_kb,
            "Sensor":       "Yes" if has_sensor else "No",
            "Temp [°C]":    round(temp_c, 1) if temp_c is not None else "—",
            "Humidity [%]": round(float(hum), 0) if hum is not None else "—",
        })

    # 3 — SoCsweep / Block bdps files (from SOC_SWEEP_RUNS already discovered)
    for info in SOC_SWEEP_RUNS:
        fn   = info["file"]
        ts   = info["ts"]
        date = ts[:10]
        time = ts[11:]
        path = info["path"]
        size_kb = os.path.getsize(path) // 1024
        temp_c, hum = _DATE_TEMP_CACHE.get(date, (None, None))
        sensor_fn = fn.replace("_bdps_", "_sensor_")
        has_sensor = os.path.exists(os.path.join(log_dir, sensor_fn))
        ocv = info.get("ocv_filename", "?")
        soc = info.get("soc_nominal", "?")
        rows.append({
            "Date":         date,
            "Time":         time.replace("-", ":"),
            "Filename":     fn,
            "Test type":    f"SoC sweep  OCV {ocv:.2f} V  ≈{soc:.0f} %",
            "Used in":      "SoC Sweep / Compare",
            "Size (KB)":    size_kb,
            "Sensor":       "Yes" if has_sensor else "No",
            "Temp [°C]":    round(temp_c, 1) if temp_c is not None else "—",
            "Humidity [%]": round(float(hum), 0) if hum is not None else "—",
        })

    # 4 — Degr_ bdps files (June 2026 new-protocol) — one summary row per session
    degr_sessions: dict = {}
    for fn in os.listdir(log_dir):
        m = _DEGR_BDPS_PAT.match(fn)
        if not m:
            continue
        cycle_num, direction, date, time = int(m.group(1)), m.group(2), m.group(3), m.group(4)
        key = (date, time)
        degr_sessions.setdefault(key, {"cycles": set(), "directions": set(), "total_kb": 0})
        degr_sessions[key]["cycles"].add(cycle_num)
        degr_sessions[key]["directions"].add(direction)
        degr_sessions[key]["total_kb"] += os.path.getsize(os.path.join(log_dir, fn)) // 1024

    for (date, time), info in sorted(degr_sessions.items()):
        nc  = sorted(info["cycles"])
        dirs = "+".join(sorted(info["directions"]))
        n_files = len(nc) * len(info["directions"])
        temp_c, hum = _DATE_TEMP_CACHE.get(date, (None, None))
        all_have_sensor = all(
            (str(c), d, date, time) in degr_sensor_keys
            for c in nc for d in info["directions"]
        )
        rows.append({
            "Date":         date,
            "Time":         time.replace("-", ":"),
            "Filename":     f"Degr_01-{max(nc):02d}_{dirs}_bdps_{date}_{time}.csv  ×{n_files} files",
            "Test type":    f"Degradation cycles 1–{max(nc)} ({dirs})",
            "Used in":      "Not yet integrated",
            "Size (KB)":    info["total_kb"],
            "Sensor":       "Yes" if all_have_sensor else "Partial",
            "Temp [°C]":    round(temp_c, 1) if temp_c is not None else "—",
            "Humidity [%]": round(float(hum), 0) if hum is not None else "—",
        })

    # 5 — DegradationCycle sensor files: one summary row per (date, time) session
    #     (226 files → ~13 summary rows; used only for temperature lookup)
    degrad_sessions: dict = {}
    for fn in os.listdir(log_dir):
        m = _DEGRAD_SENSOR_PAT.match(fn)
        if not m:
            continue
        cycle_num, _, date, time = int(m.group(1)), m.group(2), m.group(3), m.group(4)
        key = (date, time)
        degrad_sessions.setdefault(key, set()).add(cycle_num)

    for (date, time), cycles in sorted(degrad_sessions.items()):
        nc = len(cycles)
        temp_c, hum = _DATE_TEMP_CACHE.get(date, (None, None))
        rows.append({
            "Date":         date,
            "Time":         time.replace("-", ":"),
            "Filename":     f"DegradationCycle_01-{max(cycles):02d}_*_sensor_{date}_{time}.csv  ×{nc*2} files",
            "Test type":    f"Degradation session sensor ({nc} sub-cycles)",
            "Used in":      "Temperature lookup",
            "Size (KB)":    "—",
            "Sensor":       "Yes",
            "Temp [°C]":    round(temp_c, 1) if temp_c is not None else "—",
            "Humidity [%]": round(float(hum), 0) if hum is not None else "—",
        })

    df = pd.DataFrame(rows).sort_values(["Date", "Time", "Test type"]).reset_index(drop=True)
    return df


_INVENTORY_DF = build_file_inventory(LOG_DIR)

_inv_type_colors = {
    "Pulse test":               "#dceefb",
    "SOH test (C/5 discharge)": "#fff3cd",
    "SoC sweep":                "#d5f5e3",
    "Degr. cycle":              "#fde8d8",
    "Degradation session":      "#f0e6fa",
    "Temperature lookup":       "#f0e6fa",
}

inventory_table = pn.widgets.Tabulator(
    _INVENTORY_DF,
    pagination="local",
    page_size=50,
    height=620,
    sizing_mode="stretch_width",
    show_index=False,
    frozen_columns=["Date", "Filename"],
    header_filters=True,
    layout="fit_data_table",
    configuration={
        "columnDefaults": {"headerFilter": True},
        "columns": [
            {"field": "Date",         "width": 100, "frozen": True},
            {"field": "Time",         "width": 80},
            {"field": "Filename",     "width": 380, "frozen": True},
            {"field": "Test type",    "width": 220},
            {"field": "Used in",      "width": 220},
            {"field": "Size (KB)",    "width": 90,  "hozAlign": "right"},
            {"field": "Sensor",       "width": 70,  "hozAlign": "center"},
            {"field": "Temp [°C]",    "width": 90,  "hozAlign": "right"},
            {"field": "Humidity [%]", "width": 110, "hozAlign": "right"},
        ],
    },
)

_n_sensor = (_INVENTORY_DF["Sensor"] == "Yes").sum()
_n_total  = len(_INVENTORY_DF)

inventory_tab = pn.Column(
    pn.pane.Markdown("## File Inventory"),
    pn.pane.Markdown(
        f"**{_n_total} files** (full_charge, cycle_ and small fragments excluded). "
        f"Sensor/temperature data available for **{_n_sensor}** of {_n_total} entries. "
        "DegradationCycle sensor files are shown as one summary row per session. "
        "Use the column header fields to filter."
    ),
    inventory_table,
    sizing_mode="stretch_width",
)


# ── Test Plan tab ──────────────────────────────────────────────────────────────
def build_test_plan_df(log_dir: str) -> pd.DataFrame:
    """
    One row per test block.  A block = one set of degradation cycles + up to 2
    realistic test days (testday_run) + one SOH test.
    Blocks are anchored by SOH tests in chronological order.
    """
    # ── collect all events ───────────────────────────────────────────────────
    _TR_PAT  = re.compile(r"^testday_run_(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2}-\d{2})_bdps\.csv$")
    _C5_PAT2 = re.compile(r"^discharge_c5_(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2}-\d{2})_bdps\.csv$")
    _DC_PAT  = re.compile(r"^DegradationCycle_(\d+)_(charge|discharge)_sensor_"
                          r"(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2}-\d{2})\.csv$")
    _DG_PAT  = re.compile(r"^Degr_(\d+)_(charge|discharge)_bdps_"
                          r"(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2}-\d{2})\.csv$")

    tr_dates:  list = []        # (date, time, filename)
    soh_tests: list = []        # (date, time, ah, v_final, complete, status)
    dc_sess:   dict = {}        # (date, time) -> set of cycle ints
    dg_sess:   dict = {}        # (date, time) -> set of cycle ints

    for fn in sorted(os.listdir(log_dir)):
        m = _TR_PAT.match(fn)
        if m and m.group(1) >= "2025-12-09":
            tr_dates.append((m.group(1), m.group(2), fn))
            continue
        m = _C5_PAT2.match(fn)
        if m:
            path = os.path.join(log_dir, fn)
            size = os.path.getsize(path)
            # Skip pre-protocol (before Dec 9) and too-small files
            if m.group(1) < "2025-12-09" or size < SOH_MIN_BYTES:
                continue
            df = pd.read_csv(path)
            df.columns = [c.strip().lower() for c in df.columns]
            dis = df[df["current"] < -0.5]
            dt  = df["elapsed_s"].diff().clip(0, 10)
            ah  = float((np.abs(dis["current"]) * dt[dis.index]).sum() / 3600)
            if ah < SOH_MIN_AH:
                continue  # skip short/aborted discharges (e.g. Mar 11 4 Ah tests)
            vf   = float(dis["voltage"].iloc[-1]) if not dis.empty else None
            comp = vf is not None and vf <= SOH_CUTOFF_V
            soh_tests.append((m.group(1), m.group(2), round(ah, 2), vf, comp, "OK"))
            continue
        m = _DC_PAT.match(fn)
        if m:
            key = (m.group(3), m.group(4))
            dc_sess.setdefault(key, set()).add(int(m.group(1)))
            continue
        m = _DG_PAT.match(fn)
        if m:
            key = (m.group(3), m.group(4))
            dg_sess.setdefault(key, set()).add(int(m.group(1)))

    # Deduplicate SOH tests on the same date: keep the one with largest capacity
    by_date: dict = {}
    for entry in soh_tests:
        d = entry[0]
        existing = by_date.get(d)
        if existing is None or (entry[2] or 0) > (existing[2] or 0):
            by_date[d] = entry
    soh_tests = list(by_date.values())

    # Add the missing Feb-10 BDPS SOH test
    soh_tests.append(("2026-02-10", "17-23-21", None, None, False, "MISSING"))
    soh_tests.sort(key=lambda x: (x[0], x[1]))

    # ── assign each event to the block it belongs to ─────────────────────────
    # Blocks are determined by SOH tests; events before the first SOH → block 1,
    # events between SOH n and SOH n+1 → block n+1, etc.
    soh_dts = [pd.Timestamp(f"{d} {t.replace('-', ':')}") for d, t, *_ in soh_tests]

    def block_for(date: str, time: str) -> int:
        dt = pd.Timestamp(f"{date} {time.replace('-', ':')}")
        for i, soh_dt in enumerate(soh_dts):
            if dt <= soh_dt:
                return i  # belongs to block i (0-indexed)
        return len(soh_dts)  # after last SOH → open block

    # Assign testday_runs to blocks
    tr_by_block: dict = {}
    for d, t, fn in tr_dates:
        b = block_for(d, t)
        tr_by_block.setdefault(b, []).append((d, fn))

    # Assign DegradCycle sessions to blocks
    dc_by_block: dict = {}
    for (d, t), cycs in dc_sess.items():
        b = block_for(d, t)
        dc_by_block.setdefault(b, []).append((d, t, cycs, "DegradCycle"))
    for (d, t), cycs in dg_sess.items():
        b = block_for(d, t)
        dc_by_block.setdefault(b, []).append((d, t, cycs, "Degr_"))

    # ── build one row per block ───────────────────────────────────────────────
    rows = []
    for i, (soh_d, soh_t, ah, vf, comp, status) in enumerate(soh_tests):
        trs   = tr_by_block.get(i, [])
        dcs   = sorted(dc_by_block.get(i, []), key=lambda x: (x[0], x[1]))

        # DegradCycle summary
        all_cycs = set()
        dc_dates_str = []
        has_degr_fmt = False
        for d, t, cycs, fmt in dcs:
            all_cycs |= cycs
            dc_dates_str.append(f"{d}  {fmt} {min(cycs)}-{max(cycs)}")
            if fmt == "Degr_":
                has_degr_fmt = True
        dc_str  = "; ".join(dc_dates_str) if dc_dates_str else "—"
        n_cycs  = len(all_cycs)

        # Testday_run slots; N/A for new Degr_ protocol blocks that never had testday_run files
        missing_td_label = "N/A (Degr_ protocol)" if has_degr_fmt else "MISSING"
        td1_date = trs[0][0] if len(trs) >= 1 else None
        td1_fn   = trs[0][1] if len(trs) >= 1 else None
        td2_date = trs[1][0] if len(trs) >= 2 else None
        td2_fn   = trs[1][1] if len(trs) >= 2 else None
        extra_td = len(trs) - 2 if len(trs) > 2 else 0

        # Temperature
        td1_temp = _DATE_TEMP_CACHE.get(td1_date, (None, None))[0] if td1_date else None
        td2_temp = _DATE_TEMP_CACHE.get(td2_date, (None, None))[0] if td2_date else None
        soh_temp = _DATE_TEMP_CACHE.get(soh_d, (None, None))[0]

        def tfmt(v): return f"{v:.1f} C" if v is not None else "—"

        def soh_cell():
            if status == "MISSING":
                return "MISSING (sensor-only, BDPS lost)"
            cap_str  = f"{ah:.2f} Ah" if ah is not None else "?"
            comp_str = "" if comp else "  [incomplete]"
            return f"{soh_d}  {cap_str}{comp_str}"

        rows.append({
            "Block":              i + 1,
            "Degr. cycles":       dc_str if dc_str != "—" else "—",
            "# sub-cycles":       n_cycs if n_cycs else "—",
            "Testday run 1":      f"{td1_date}  {td1_fn}" if td1_fn else missing_td_label,
            "Temp TD1 [C]":       tfmt(td1_temp),
            "Testday run 2":      f"{td2_date}  {td2_fn}" if td2_fn else missing_td_label,
            "Temp TD2 [C]":       tfmt(td2_temp),
            "Extra runs":         extra_td if extra_td else "—",
            "SOH test":           soh_cell(),
            "Capacity [Ah]":      round(ah, 2) if ah is not None else "—",
            "Complete":           "Yes" if comp else ("MISSING" if status == "MISSING" else "No"),
            "Temp SOH [C]":       tfmt(soh_temp),
        })

    # Open block (after last SOH)
    i = len(soh_tests)
    trs = tr_by_block.get(i, [])
    dcs = sorted(dc_by_block.get(i, []), key=lambda x: (x[0], x[1]))
    all_cycs = set()
    dc_dates_str = []
    has_degr_open = False
    for d, t, cycs, fmt in dcs:
        all_cycs |= cycs
        dc_dates_str.append(f"{d}  {fmt} {min(cycs)}-{max(cycs)}")
        if fmt == "Degr_":
            has_degr_open = True
    dc_str   = "; ".join(dc_dates_str) if dc_dates_str else "—"
    miss_lbl = "N/A (Degr_ protocol)" if has_degr_open else "MISSING"
    td1_d    = trs[0][0] if len(trs) >= 1 else None
    td1      = trs[0][1] if len(trs) >= 1 else None
    td2_d    = trs[1][0] if len(trs) >= 2 else None
    td2      = trs[1][1] if len(trs) >= 2 else None
    rows.append({
        "Block":          i + 1,
        "Degr. cycles":   dc_str,
        "# sub-cycles":   len(all_cycs) if all_cycs else "—",
        "Testday run 1":  f"{td1_d}  {td1}" if td1 else miss_lbl,
        "Temp TD1 [C]":   "—",
        "Testday run 2":  f"{td2_d}  {td2}" if td2 else miss_lbl,
        "Temp TD2 [C]":   "—",
        "Extra runs":     max(0, len(trs) - 2) or "—",
        "SOH test":       "NOT YET DONE",
        "Capacity [Ah]":  "—",
        "Complete":       "—",
        "Temp SOH [C]":   "—",
    })

    return pd.DataFrame(rows)


_TEST_PLAN_DF = build_test_plan_df(LOG_DIR)

_PLAN_STATUS_CSS = """
.tabulator-row[tabulator-row-index] .tabulator-cell[tabulator-field='Testday run 2'] { }
"""

def _plan_row_style(row):
    styles = {}
    if str(row.get("Testday run 2", "")).startswith("MISSING"):
        styles["Testday run 2"] = "background-color: #ffe0e0; color: #cc0000; font-weight: bold"
    if str(row.get("Testday run 1", "")).startswith("MISSING"):
        styles["Testday run 1"] = "background-color: #ffe0e0; color: #cc0000; font-weight: bold"
    if str(row.get("SOH test", "")).startswith("MISSING") or str(row.get("SOH test", "")).startswith("NOT"):
        styles["SOH test"]      = "background-color: #ffe0e0; color: #cc0000; font-weight: bold"
    return styles

plan_table = pn.widgets.Tabulator(
    _TEST_PLAN_DF,
    pagination=None,
    height=700,
    sizing_mode="stretch_width",
    show_index=False,
    frozen_columns=["Block"],
    layout="fit_data_table",
    configuration={
        "columns": [
            {"field": "Block",         "width": 55,  "hozAlign": "center", "frozen": True},
            {"field": "Degr. cycles",  "width": 340},
            {"field": "# sub-cycles",  "width": 100, "hozAlign": "center"},
            {"field": "Testday run 1", "width": 380},
            {"field": "Temp TD1 [C]",  "width": 100, "hozAlign": "right"},
            {"field": "Testday run 2", "width": 380},
            {"field": "Temp TD2 [C]",  "width": 100, "hozAlign": "right"},
            {"field": "Extra runs",    "width": 90,  "hozAlign": "center"},
            {"field": "SOH test",      "width": 380},
            {"field": "Capacity [Ah]", "width": 110, "hozAlign": "right"},
            {"field": "Complete",      "width": 90,  "hozAlign": "center"},
            {"field": "Temp SOH [C]",  "width": 110, "hozAlign": "right"},
        ],
    },
)

_plan_missing_td2 = (_TEST_PLAN_DF["Testday run 2"].str.startswith("MISSING")).sum()
_plan_missing_soh = (_TEST_PLAN_DF["SOH test"].str.contains("MISSING|NOT YET", na=False)).sum()
_plan_total_blk   = len(_TEST_PLAN_DF) - 1  # exclude open block

plan_tab = pn.Column(
    pn.pane.Markdown("## Test Plan Overview"),
    pn.pane.Markdown(
        f"**{_plan_total_blk} completed blocks** (plus 1 open block in progress).  "
        "Each block = 10 degradation cycles + **2 realistic test days** (testday_run) "
        "+ 1 SOH test (C/5 full discharge).  \n"
        f"**{_plan_missing_td2}** blocks are missing the second testday_run.  "
        f"**{_plan_missing_soh}** SOH tests are missing or not yet done.  "
        "June 2026 blocks use the new **Degr_** protocol which does not produce testday_run files — "
        "those cells show *N/A (Degr_ protocol)*."
    ),
    plan_table,
    sizing_mode="stretch_width",
)


dashboard = pn.Tabs(
    ("Run Inspector",       inspector_tab),
    ("Compare Runs",        compare_tab),
    ("SOH History",         soh_tab),
    ("Degradation Trends",  trends_tab),
    ("SoC Sweep",           sweep_tab),
    ("Test-Day v2 (Beta)",  v2_tab),
    ("Feature Correlations",corr_tab),
    ("File Inventory",      inventory_tab),
    ("Test Plan",           plan_tab),
    ("Export",              export_tab),
    sizing_mode="stretch_both",
)

dashboard.servable()

if __name__ == "__main__":
    pn.serve(dashboard, title="Battery Feature Dashboard", show=True)
