"""
battery_feature_dashboard_unit2.py

Dashboard for the new test battery **ul18_12_unit2**, logged in the v2
event-scripted test-day profile format from day one (see CLAUDE.md →
"Test-Day Profile Format (v2)" and testday_v2_features.py, which does the
actual feature extraction here).

Companion to battery_feature_dashboard.py (the original "old_ul18_12" / mostly
pre-v2 dashboard) — kept as a separate script so the two batteries are never
pooled together, per CLAUDE.md's Battery_ID rule.

As of this writing only Block 1 exists and testing is still in progress
(9 of 10 degradation cycles logged so far) — every tab here is written to
handle partial data (empty/short/fragment files) without crashing, and grows
automatically as more blocks and cycles land. Use the **Rescan** button to
pick up new files without restarting the server.

Run with:
    panel serve battery_feature_dashboard_unit2.py --show
"""

import os
import re
import warnings

import numpy as np
import pandas as pd
import panel as pn
import plotly.graph_objects as go
from plotly.subplots import make_subplots

import testday_v2_features as v2feat

warnings.filterwarnings("ignore")
pn.extension("plotly", "tabulator")

# ── Config ────────────────────────────────────────────────────────────────────
BASE_DIR       = os.path.dirname(__file__)
LOG_DIR        = os.path.join(BASE_DIR, "ul18_12_unit2")
BATTERY_ID     = "ul18_12_unit2"
NOMINAL_AH     = 18.0     # Ultracell UL18-12 nameplate C/20 capacity
MIN_DEGR_BYTES = 5_000    # Degr charge/discharge files smaller than this are fragments/aborted

# Validated categorical palette (fixed order — CVD-safe adjacent pairs; the
# first 3 slots also clear all-pairs comparisons, so scatter/bubble plots
# should stay within them where practical). Status colors are reserved and
# never reused as series colors, so state (complete/in-progress/invalid) is
# always visually distinct from series identity (which run, which feature).
_COLORS = [
    "#2a78d6",  # 1 blue
    "#eb6834",  # 2 orange
    "#1baf7a",  # 3 aqua
    "#eda100",  # 4 yellow
    "#e87ba4",  # 5 magenta
    "#008300",  # 6 green
    "#4a3aa7",  # 7 violet
    "#e34948",  # 8 red
]
_STATUS_GOOD     = "#0ca30c"
_STATUS_WARNING  = "#fab219"
_STATUS_SERIOUS  = "#ec835a"
_STATUS_CRITICAL = "#d03b3b"
_DIVERGING_LO, _DIVERGING_MID, _DIVERGING_HI = "#2a78d6", "#f0efec", "#e34948"
_DIVERGING_SCALE = [[0.0, _DIVERGING_LO], [0.5, _DIVERGING_MID], [1.0, _DIVERGING_HI]]
# Sequential = one hue, light -> dark (same blue as categorical slot 1)
_SEQUENTIAL_BLUE = [
    [0.0, "#cde2fb"], [0.25, "#86b6ef"], [0.5, "#3987e5"],
    [0.75, "#1c5cab"], [1.0, "#0d366b"],
]
_EVENT_COLORS = {
    "rest_baseline":       "#e0e0e0",
    "ocv_window":          "#c7c7c7",
    "wakeup_load":         "#ffbb78",
    "glow_plug_like_load": "#ff9896",
    "crank_pulse":         "#d62728",
    "recovery_rest":       "#aec7e8",
    "alternator_charge":   "#98df8a",
    "ramp_like_load":      "#c5b0d5",
    "driving_aux_load":    "#9edae5",
}
_META_COLS = {"battery_id", "block", "kind", "timestamp", "filename", "ocv_label_V"}

# OCV -> SoC lookup for the Ultracell UL18-12 VRLA/AGM at 25°C (same table as
# battery_feature_dashboard.py) — used to show SoC alongside OCV in hovers.
_OCV_V   = np.array([10.50, 11.51, 11.66, 11.81, 11.96, 12.10, 12.20, 12.32, 12.42, 12.50, 12.70])
_SOC_PCT = np.array([   0,    10,    20,    30,    40,    50,    60,    70,    80,    90,   100])


def _v_to_soc(v_ocv: float) -> float | None:
    if v_ocv is None or pd.isna(v_ocv):
        return None
    return round(float(np.interp(v_ocv, _OCV_V, _SOC_PCT)), 0)

_DEGR_PAT = re.compile(
    r"^Block_(\d+)_Degr_(\d+)_(charge|discharge)_bdps_"
    r"(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})\.csv$"
)


# ── Degradation-cycle discovery (Block_<nn>_Degr_<cycle>_charge/discharge) ────

SAMPLE_GAP_COARSE_S = 5.0  # median inter-sample gap above this -> dt-clip integration undercounts


def _cycle_ah(path: str, sign: str) -> tuple:
    """Ah + completeness + coarse-sampling flag for one charge/discharge Degr
    file.

    Fragments (e.g. an aborted cycle logged as a 3-row/246-byte file) are
    flagged, not silently dropped, since incomplete data is the expected
    state here. Separately: starting mid-Block 4 the Pi's logging interval
    degraded badly (median gap up to ~45s by Block 5, against a 10s dt-clip)
    — Degr *discharge* current is confirmed constant (~-9A) in this test
    protocol, so it's corrected the same way as the SOH C/5 calc
    (duration x mean|I|, immune to sample sparsity). Degr *charge* is CC/CV
    (current tapers from ~5A to <1A near the end) — a coarse-sampled charge
    file has no reliable correction, so it's just flagged unreliable rather
    than "fixed" with a number that can't be trusted.

    Returns (ah, n_rows, size_bytes, complete, coarse_sampling).
    """
    size = os.path.getsize(path)
    df = pd.read_csv(path)
    n = len(df)
    if size < MIN_DEGR_BYTES or n < 10:
        return None, n, size, False, False

    gap_series = df["Elapsed_s"].diff()
    median_gap = gap_series.median()
    coarse = bool(pd.notna(median_gap) and median_gap > SAMPLE_GAP_COARSE_S)

    if sign == "discharge":
        mask = df["Current"] < -0.5
        if coarse:
            duration_s = df["Elapsed_s"].max() - df["Elapsed_s"].min()
            ah = float(df.loc[mask, "Current"].abs().mean() * duration_s / 3600)
        else:
            dt = gap_series.clip(0, 10)
            ah = float((df.loc[mask, "Current"].abs() * dt.loc[mask]).sum() / 3600)
    else:
        mask = df["Current"] > 0.5
        dt = gap_series.clip(0, 10)
        ah = float((df.loc[mask, "Current"] * dt.loc[mask]).sum() / 3600)

    return ah, n, size, True, coarse


def find_degr_cycles(log_dir: str) -> pd.DataFrame:
    """One row per (block, cycle): discharge/charge Ah, coulombic efficiency,
    and a completeness flag. Ready for as many blocks/cycles as show up.

    A cycle can look well-formed (passes the size/row-count check) but still
    be a truncated in-progress write — e.g. a charge cycle logged mid-charge
    has a normal header and thousands of rows, just far less Ah than a
    finished one. So completeness is refined in a second pass: any Ah value
    below half the block's own median for that charge/discharge is flagged
    too, rather than trusting the file-size check alone.

    A given (block, cycle, charge/discharge) can have more than one file on
    disk — e.g. a coarse-sampled cycle logged during the logging-rate
    regression, later re-run cleanly. `ts` sorts lexicographically the same
    as chronologically, so only the most recent file per identity is kept.
    """
    entries: dict = {}
    ts_seen: dict = {}
    for fn in sorted(os.listdir(log_dir)):
        m = _DEGR_PAT.match(fn)
        if not m:
            continue
        block, cycle, kind, ts = int(m.group(1)), int(m.group(2)), m.group(3), m.group(4)
        key = (block, cycle, kind)
        if key not in ts_seen or ts > ts_seen[key]:
            ts_seen[key] = ts
            entries.setdefault((block, cycle), {})[kind] = os.path.join(log_dir, fn)

    rows = []
    for (block, cycle), kinds in sorted(entries.items()):
        row = {"block": block, "cycle": cycle}
        for kind in ("discharge", "charge"):
            if kind in kinds:
                ah, n_rows, size, complete, coarse = _cycle_ah(kinds[kind], kind)
                row[f"{kind}_ah"] = ah
                row[f"{kind}_complete"] = complete
                row[f"{kind}_coarse_sampling"] = coarse
            else:
                row[f"{kind}_ah"] = None
                row[f"{kind}_complete"] = False
                row[f"{kind}_coarse_sampling"] = False
        rows.append(row)

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    for kind in ("discharge", "charge"):
        col = f"{kind}_ah"
        # Per-block median (this battery's Ah scale shrinks a lot over 20
        # blocks of degradation — a global median would flag every cycle in
        # a heavily-degraded late block as "too low" relative to the early,
        # healthy blocks' much larger Ah values).
        valid_ah = df[col].where(df[f"{kind}_complete"])
        median_by_block = valid_ah.groupby(df["block"]).transform("median")
        too_low = (
            df[col].notna() & median_by_block.notna() & (median_by_block > 0)
            & (df[col] < 0.5 * median_by_block)
        )
        df.loc[too_low, f"{kind}_complete"] = False

    df["coulombic_eff_pct"] = None
    both = df["discharge_ah"].notna() & df["charge_ah"].notna()
    df.loc[both, "coulombic_eff_pct"] = df.loc[both, "charge_ah"] / df.loc[both, "discharge_ah"] * 100

    def _status(r):
        if not (r["discharge_complete"] and r["charge_complete"]):
            return "in progress / fragment"
        if r["discharge_coarse_sampling"] or r["charge_coarse_sampling"]:
            return "complete (Ah corrected/uncertain — coarse Pi sampling)"
        return "complete"

    df["status"] = df.apply(_status, axis=1)
    return df


# ── File inventory (all file kinds, for tracking what's landed so far) ───────

def build_file_inventory(log_dir: str) -> pd.DataFrame:
    rows = []
    for fn in sorted(os.listdir(log_dir)):
        if not fn.endswith(".csv") or "_sensor_" in fn or "_sensor.csv" in fn:
            continue  # list the bdps side only; sensor files are paired 1:1
        path = os.path.join(log_dir, fn)
        if "_SOH_C5_" in fn:
            kind = "SOH C/5"
        elif "_testday_" in fn:
            kind = "Test-day (v2)"
        elif "_stepdown_" in fn:
            kind = "Step-down"
        elif "_Degr_" in fn and "_charge_" in fn:
            kind = "Degr charge"
        elif "_Degr_" in fn and "_discharge_" in fn:
            kind = "Degr discharge"
        elif "charge_full" in fn:
            kind = "Initial full charge"
        elif "sweep_charge" in fn:
            kind = "Sweep recharge"
        else:
            kind = "other"
        m_block = re.search(r"Block_(\d+)", fn)
        m_cycle = re.search(r"Degr_(\d+)", fn)
        rows.append({
            "file": fn,
            "block": int(m_block.group(1)) if m_block else None,
            "cycle": int(m_cycle.group(1)) if m_cycle else None,
            "kind": kind,
            "size_kb": round(os.path.getsize(path) / 1024, 1),
        })
    return pd.DataFrame(rows)


# ── Run Inspector figure (event-shaded V/I traces) ───────────────────────────

def build_event_fig(bdps_path: str, sensor_path: str | None = None) -> go.Figure:
    df = v2feat.load_run(bdps_path, sensor_path)
    if df is None:
        fig = go.Figure()
        fig.update_layout(title="This file does not have the v2 event schema.", height=300)
        return fig

    fig = make_subplots(
        rows=2, cols=1, shared_xaxes=True, vertical_spacing=0.08,
        subplot_titles=["Voltage", "Current"], row_heights=[0.55, 0.45],
    )
    seen_types = []
    for _, g in df.groupby("Event_Index", sort=True):
        etype = g["Event_Type"].iloc[0]
        color = _EVENT_COLORS.get(etype, "#eeeeee")
        t0, t1 = g["Elapsed_s"].min(), g["Elapsed_s"].max()
        for row in (1, 2):
            fig.add_vrect(x0=t0, x1=t1, fillcolor=color, opacity=0.25, line_width=0, row=row, col=1)
        if etype not in seen_types:
            seen_types.append(etype)

    fig.add_trace(go.Scatter(
        x=df["Elapsed_s"], y=df["Voltage"], mode="lines",
        name="Voltage", line=dict(color=_COLORS[0], width=1.5),
    ), row=1, col=1)
    fig.add_trace(go.Scatter(
        x=df["Elapsed_s"], y=df["Current"], mode="lines",
        name="Current", line=dict(color=_COLORS[7], width=1.5), showlegend=False,
    ), row=2, col=1)

    for etype in seen_types:
        fig.add_trace(go.Scatter(
            x=[None], y=[None], mode="markers",
            marker=dict(size=10, color=_EVENT_COLORS.get(etype, "#eeeeee")),
            name=etype,
        ), row=1, col=1)

    fig.update_yaxes(title_text="Voltage [V]", row=1, col=1)
    fig.update_yaxes(title_text="Current [A]", row=2, col=1)
    fig.update_xaxes(title_text="Elapsed [s]", row=2, col=1)
    fig.update_layout(
        height=580, template="plotly_white",
        legend=dict(orientation="h", y=-0.15),
        margin=dict(t=50, b=40),
    )
    return fig


def build_feature_table(bdps_path: str, sensor_path: str | None = None) -> pd.DataFrame:
    df = v2feat.load_run(bdps_path, sensor_path)
    if df is None:
        return pd.DataFrame(columns=["feature", "value"])
    feats = v2feat.extract_features(df)
    return pd.DataFrame(sorted(feats.items()), columns=["feature", "value"])


# ── Compare Runs figure ────────────────────────────────────────────────────────

def build_compare_fig(labels: list, run_options: dict) -> go.Figure:
    fig = make_subplots(
        rows=2, cols=1, shared_xaxes=True, vertical_spacing=0.08,
        subplot_titles=["Voltage", "Current"], row_heights=[0.55, 0.45],
    )
    for i, label in enumerate(labels):
        info = run_options.get(label)
        if info is None:
            continue
        df = v2feat.load_run(info["bdps_path"])
        if df is None:
            continue
        color = _COLORS[i % len(_COLORS)]
        name = f"OCV {info['ocv_label']:.2f}V" if info["ocv_label"] is not None else info["filename"]
        fig.add_trace(go.Scatter(x=df["Elapsed_s"], y=df["Voltage"], mode="lines",
                                  name=name, line=dict(color=color)), row=1, col=1)
        fig.add_trace(go.Scatter(x=df["Elapsed_s"], y=df["Current"], mode="lines",
                                  showlegend=False, line=dict(color=color)), row=2, col=1)
    fig.update_yaxes(title_text="Voltage [V]", row=1, col=1)
    fig.update_yaxes(title_text="Current [A]", row=2, col=1)
    fig.update_xaxes(title_text="Elapsed [s]", row=2, col=1)
    fig.update_layout(height=580, template="plotly_white", legend=dict(orientation="h", y=-0.15))
    return fig


# ── Feature Correlations ───────────────────────────────────────────────────────
_CORR_EXCLUDE_COLS = {"battery_id", "kind", "timestamp", "filename"}


def _corr_columns(df: pd.DataFrame) -> list:
    """Numeric, non-identifier columns with actual variance in df. A
    constant column (e.g. n_events/cutoff_hit are the same value for every
    row once filtered to "clean" runs) has an undefined (NaN) correlation
    with everything, including itself — dropped here rather than left to
    show up as blank NaN rows/columns in the heatmap."""
    return [
        c for c in df.columns
        if c not in _CORR_EXCLUDE_COLS
        and pd.api.types.is_numeric_dtype(df[c])
        and df[c].nunique(dropna=True) > 1
    ]


def build_run_corr_heatmap(df: pd.DataFrame, cols: list) -> go.Figure:
    """Spearman correlation matrix across run-level test-day features."""
    available = [
        c for c in cols
        if c in df.columns and df[c].notna().sum() > 3 and df[c].nunique(dropna=True) > 1
    ]
    if len(available) < 2:
        return go.Figure().update_layout(title="Not enough data yet", height=520)
    sub = df[available].apply(pd.to_numeric, errors="coerce")
    corr = sub.corr(method="spearman").round(2)
    z = corr.values
    labels = corr.columns.tolist()
    # A pair can still be individually undefined even after dropping globally
    # constant columns — e.g. in "include all runs" mode, Block 5's corrupted
    # event structure means a column like crank_cold_V_pre is only populated
    # for clean rows, so correlating it against cutoff_hit hits a subset
    # where cutoff_hit has no variance. Genuinely undefined, not a bug —
    # shown as a blank cell rather than the literal text "nan".
    text = [[("" if pd.isna(v) else f"{v:.2f}") for v in row] for row in z]
    hover_r = [[("not enough overlapping data" if pd.isna(v) else f"{v:.3f}") for v in row] for row in z]
    fig = go.Figure(go.Heatmap(
        z=z, x=labels, y=labels, text=text, texttemplate="%{text}",
        textfont=dict(size=9),
        colorscale=_DIVERGING_SCALE, zmid=0, zmin=-1, zmax=1,
        colorbar=dict(title="Spearman r", thickness=14, len=0.9),
        customdata=hover_r,
        hovertemplate="<b>%{y}</b> vs <b>%{x}</b><br>r = %{customdata}<extra></extra>",
        xgap=1, ygap=1,
    ))
    fig.update_layout(
        title=dict(text="Spearman correlation matrix — test-day run features", x=0.02, xanchor="left"),
        template="plotly_white",
        height=max(560, 22 * len(labels)), margin=dict(l=170, r=60, t=50, b=170),
        xaxis=dict(tickangle=-40), yaxis=dict(autorange="reversed"),
    )
    return fig


def build_scatter_pair(df: pd.DataFrame, x_col: str, y_col: str, color_col: str = "block") -> go.Figure:
    """Scatter of two features, coloured by a third (block by default).
    Every point's hover shows the source filename, block, starting OCV and
    SoC% — the run identity behind that dot — regardless of which columns
    are plotted or used for colour."""
    if x_col not in df.columns or y_col not in df.columns:
        return go.Figure().update_layout(title="Column not found")

    id_cols = ["filename", "block", "ocv_label_V", "soc_pct"]
    want_cols = list(dict.fromkeys(
        [x_col, y_col] + [c for c in id_cols if c in df.columns]
        + ([color_col] if color_col in df.columns else [])
    ))
    sub = df[want_cols].dropna(subset=[x_col, y_col])
    if sub.empty:
        return go.Figure().update_layout(title="No data")

    fname = sub["filename"] if "filename" in sub.columns else pd.Series("n/a", index=sub.index)
    block = sub["block"] if "block" in sub.columns else pd.Series(np.nan, index=sub.index)
    ocv = sub["ocv_label_V"] if "ocv_label_V" in sub.columns else pd.Series(np.nan, index=sub.index)
    soc = sub["soc_pct"] if "soc_pct" in sub.columns else pd.Series(np.nan, index=sub.index)
    customdata = np.stack([
        fname.astype(str).values,
        block.map(lambda v: "n/a" if pd.isna(v) else f"{int(v)}").values,
        ocv.map(lambda v: "n/a" if pd.isna(v) else f"{v:.2f}").values,
        soc.map(lambda v: "n/a" if pd.isna(v) else f"{v:.0f}").values,
    ], axis=-1)
    id_line = (
        "<b>%{customdata[0]}</b><br>"
        "Block %{customdata[1]}  ·  OCV %{customdata[2]} V  ·  SoC ≈ %{customdata[3]}%<br>"
    )

    if color_col in sub.columns:
        marker = dict(
            color=pd.to_numeric(sub[color_col], errors="coerce"), colorscale=_SEQUENTIAL_BLUE, showscale=True,
            colorbar=dict(title=color_col, thickness=14, len=0.75),
            size=10, opacity=0.85, line=dict(width=1, color="white"),
        )
        htmpl = id_line + f"{x_col} = %{{x}}<br>{y_col} = %{{y}}<br>{color_col} = %{{marker.color}}<extra></extra>"
    else:
        marker = dict(size=10, color=_COLORS[0], opacity=0.85, line=dict(width=1, color="white"))
        htmpl = id_line + f"{x_col} = %{{x}}<br>{y_col} = %{{y}}<extra></extra>"

    x_n = pd.to_numeric(sub[x_col], errors="coerce")
    y_n = pd.to_numeric(sub[y_col], errors="coerce")
    valid = np.isfinite(x_n) & np.isfinite(y_n)
    traces = [go.Scatter(
        x=sub[x_col], y=sub[y_col], mode="markers", marker=marker,
        customdata=customdata, hovertemplate=htmpl, name="data",
    )]
    if valid.sum() >= 3:
        m, b = np.polyfit(x_n[valid], y_n[valid], 1)
        x_fit = np.linspace(x_n[valid].min(), x_n[valid].max(), 100)
        r = np.corrcoef(x_n[valid], y_n[valid])[0, 1]
        traces.append(go.Scatter(
            x=x_fit, y=m * x_fit + b, mode="lines",
            line=dict(color=_COLORS[7], dash="dash", width=1.5),
            name=f"trend  r={r:.2f}", showlegend=True,
        ))
    fig = go.Figure(traces)
    fig.update_layout(
        title=dict(text=f"{y_col}  vs  {x_col}", x=0.02, xanchor="left"),
        xaxis_title=x_col, yaxis_title=y_col,
        template="plotly_white", height=460, margin=dict(l=60, r=40, t=60, b=50),
        legend=dict(orientation="h", y=1.1),
    )
    return fig


def build_block_predictor_table(dataset: pd.DataFrame, soh_df: pd.DataFrame) -> pd.DataFrame:
    """For each within-block test-day feature (mean over that block's clean,
    non-coarse-sampled runs), Spearman r against THIS block's own measured
    C/5 capacity and against the NEXT block's capacity. This is the core
    research question the v2 profile was designed for (see
    testday_v2_features.py's module docstring): the vehicle only ever
    observes OCV, never true SOH — a feature that predicts the *next*
    block's capacity ahead of time is a candidate field-deployable health
    indicator. Very small sample (one row per block) — exploratory only.

    soh_df comes from find_block_soh_df() rather than dataset's own
    block_c5_capacity_ah column, since SOH is measured before that block's
    test-day runs exist — a block with a fresh SOH reading but no test-day
    runs yet (e.g. right after `charge_full`) is still a valid "next block"
    target for the previous block's predictors, even though it has no rows
    in `dataset` at all.
    """
    cols = ["feature", "r_this_block", "n_this", "r_next_block", "n_next"]
    if dataset.empty or "cutoff_hit" not in dataset.columns or soh_df.empty:
        return pd.DataFrame(columns=cols)

    clean = dataset[~dataset["cutoff_hit"]]
    feat_cols = [
        c for c in _corr_columns(clean)
        if c not in {"block", "ocv_label_V", "soc_pct", "block_c5_capacity_ah",
                     "block_c5_reached_cutoff", "n_events", "cutoff_hit"}
    ]
    if clean.empty or not feat_cols:
        return pd.DataFrame(columns=cols)

    per_block = clean.groupby("block")[feat_cols].mean()
    valid_soh = soh_df[soh_df["reached_cutoff"]] if "reached_cutoff" in soh_df.columns else soh_df
    soh = valid_soh.drop_duplicates(subset=["block"]).set_index("block")["block_c5_capacity_ah"].sort_index()
    this_capacity = soh.reindex(per_block.index)
    next_capacity = pd.Series(soh.reindex(per_block.index + 1).values, index=per_block.index)

    rows = []
    for feat in feat_cols:
        x = per_block[feat]
        m_this = x.notna() & this_capacity.notna()
        m_next = x.notna() & next_capacity.notna()
        r_this = x[m_this].corr(this_capacity[m_this], method="spearman") if m_this.sum() >= 3 else None
        r_next = x[m_next].corr(next_capacity[m_next], method="spearman") if m_next.sum() >= 3 else None
        if r_this is None and r_next is None:
            continue
        rows.append({
            "feature": feat,
            "r_this_block": round(r_this, 3) if r_this is not None else None,
            "n_this": int(m_this.sum()),
            "r_next_block": round(r_next, 3) if r_next is not None else None,
            "n_next": int(m_next.sum()),
        })
    out = pd.DataFrame(rows, columns=cols)
    if not out.empty:
        out = out.reindex(out["r_next_block"].abs().sort_values(ascending=False, na_position="last").index)
        out = out.reset_index(drop=True)
    return out


def build_soc_robust_indicator_table(dataset: pd.DataFrame, soh_df: pd.DataFrame) -> pd.DataFrame:
    """Rank features by how well they track SOH *without* just being a proxy
    for SoC — the property a real field-deployable indicator needs, since a
    vehicle can't control what SoC it happens to measure at.

    For each feature: `r_soc` = Spearman r against soc_pct across all clean
    runs (SoC set-points repeat every block, so this isolates genuine
    SoC-dependence rather than a block/degradation confound); `r_next_block`
    = Spearman r between that block's mean feature value and the *next*
    block's measured capacity (from build_block_predictor_table). `quality`
    = |r_next_block| − |r_soc|: high when a feature strongly tracks future
    SOH while barely moving with SoC. Absolute voltage-level features
    (OCV/rest/recovery/crank V_pre, …) score poorly here almost by
    construction — voltage directly encodes charge state for a lead-acid
    cell, so they're highly SoC-coupled regardless of degradation.
    Resistance-derived features (crank/wakeup/ramp R_int) are the physically
    expected winners: they reflect internal impedance, not charge level.
    """
    cols = ["feature", "r_soc", "r_this_block", "r_next_block", "n_next", "quality"]
    predictors = build_block_predictor_table(dataset, soh_df)
    if predictors.empty or dataset.empty or "cutoff_hit" not in dataset.columns:
        return pd.DataFrame(columns=cols)

    clean = dataset[~dataset["cutoff_hit"]]
    if "soc_pct" not in clean.columns:
        return pd.DataFrame(columns=cols)

    r_soc = {}
    for feat in predictors["feature"]:
        sub = clean[[feat, "soc_pct"]].dropna()
        if len(sub) >= 5:
            r_soc[feat] = sub[feat].corr(sub["soc_pct"], method="spearman")

    out = predictors[predictors["feature"].isin(r_soc)].copy()
    out["r_soc"] = out["feature"].map(r_soc).round(3)
    out["quality"] = (out["r_next_block"].abs() - out["r_soc"].abs()).round(3)
    out = out[cols].sort_values("quality", ascending=False).reset_index(drop=True)
    return out


# ── SoC Sweep / Degradation Trends / SOH figures ──────────────────────────────

def _feature_options(df: pd.DataFrame) -> list:
    if df.empty:
        return []
    return [
        c for c in df.columns
        if c not in _META_COLS and pd.api.types.is_numeric_dtype(df[c])
    ]


def build_sweep_fig(df: pd.DataFrame, feature: str) -> go.Figure:
    fig = go.Figure()
    if df.empty or feature not in df.columns or df[feature].dropna().empty:
        fig.update_layout(title="No data for this feature yet", height=420, template="plotly_white")
        return fig
    sub = df.dropna(subset=[feature]).sort_values("ocv_label_V")
    fig.add_trace(go.Scatter(
        x=sub["ocv_label_V"], y=sub[feature], mode="markers+lines",
        marker=dict(size=9, color=_COLORS[0]),
    ))
    fig.update_xaxes(title_text="Starting OCV [V]")
    fig.update_yaxes(title_text=feature)
    fig.update_layout(height=420, template="plotly_white", title=f"{feature} vs starting OCV")
    return fig


def build_trend_fig(df: pd.DataFrame, feature: str) -> go.Figure:
    fig = go.Figure()
    if df.empty or feature not in df.columns or df[feature].dropna().empty:
        fig.update_layout(title="No data for this feature yet", height=420, template="plotly_white")
        return fig
    sub = df.dropna(subset=[feature])
    agg = sub.groupby("block")[feature].mean().reset_index()
    fig.add_trace(go.Scatter(
        x=agg["block"], y=agg[feature], mode="markers+lines",
        marker=dict(size=10, color=_COLORS[0]),
    ))
    n_blocks = df["block"].nunique()
    note = "" if n_blocks > 1 else "  (single block so far — trend will develop as more blocks are tested)"
    fig.update_xaxes(title_text="Block #", dtick=1)
    fig.update_yaxes(title_text=feature)
    fig.update_layout(height=420, template="plotly_white", title=f"{feature} vs block{note}")
    return fig


def build_soh_fig(soh: pd.DataFrame, nominal_ah: float = NOMINAL_AH) -> go.Figure:
    """soh: tidy (block, block_c5_capacity_ah, reached_cutoff) frame from
    testday_v2_features.find_block_soh_df() — independent of whether that
    block has any test-day runs yet, since SOH is measured before the
    SoC-sweep starts.

    A discharge that hasn't reached the real cutoff voltage yet most likely
    means the C/5 test was still running on the Pi at the moment this file
    was synced (e.g. Block 6 read 12.07 V at sync time — every *completed*
    discharge so far ends at ~10.78-10.79 V). Its Ah is just a snapshot
    lower bound, not the final SOH reading, so it's plotted as a separate,
    disconnected series rather than joined into the trend line — rescan
    once it finishes to get the real value.
    """
    fig = go.Figure()
    if soh.empty:
        fig.update_layout(title="No SOH (C/5) measurement yet", height=420, template="plotly_white")
        return fig
    valid = soh[soh["reached_cutoff"]] if "reached_cutoff" in soh.columns else soh
    invalid = soh[~soh["reached_cutoff"]] if "reached_cutoff" in soh.columns else soh.iloc[0:0]
    fig.add_trace(go.Scatter(
        x=valid["block"], y=valid["block_c5_capacity_ah"], mode="markers+lines",
        name="Measured C/5 Ah", line=dict(color=_COLORS[0]), marker=dict(size=10, color=_COLORS[0]),
    ))
    if not invalid.empty:
        fig.add_trace(go.Scatter(
            x=invalid["block"], y=invalid["block_c5_capacity_ah"], mode="markers",
            name="Not final yet — test likely still running (snapshot)",
            marker=dict(size=13, color=_STATUS_WARNING, symbol="x"),
        ))
    fig.add_hline(
        y=nominal_ah, line=dict(color="#898781", dash="dot"),
        annotation_text=f"Nameplate {nominal_ah:.0f} Ah", annotation_position="top left",
    )
    fig.update_xaxes(title_text="Block #", dtick=1)
    fig.update_yaxes(title_text="C/5 capacity [Ah]")
    fig.update_layout(
        height=420, template="plotly_white",
        title="SOH History — measured C/5 discharge capacity per block",
        legend=dict(orientation="h", y=-0.18),
    )
    return fig


def build_degr_fig(degr_df: pd.DataFrame) -> go.Figure:
    fig = make_subplots(rows=1, cols=2, subplot_titles=["Ah per degradation cycle", "Coulombic efficiency"])
    if degr_df.empty:
        fig.update_layout(height=420, template="plotly_white", title="No degradation-cycle data yet")
        return fig
    is_complete = degr_df["status"].str.startswith("complete")
    complete = degr_df[is_complete]
    partial = degr_df[~is_complete]

    fig.add_trace(go.Scatter(x=complete["cycle"], y=complete["discharge_ah"], mode="markers+lines",
                              name="Discharge Ah", marker=dict(color=_COLORS[0])), row=1, col=1)
    fig.add_trace(go.Scatter(x=complete["cycle"], y=complete["charge_ah"], mode="markers+lines",
                              name="Charge Ah", marker=dict(color=_COLORS[1])), row=1, col=1)
    if not partial.empty:
        fig.add_trace(go.Scatter(x=partial["cycle"], y=partial["discharge_ah"], mode="markers",
                                  name="In progress / fragment",
                                  marker=dict(color=_STATUS_WARNING, symbol="x", size=10)), row=1, col=1)
        fig.add_trace(go.Scatter(x=partial["cycle"], y=partial["charge_ah"], mode="markers",
                                  showlegend=False,
                                  marker=dict(color=_STATUS_WARNING, symbol="x", size=10)), row=1, col=1)
    fig.add_trace(go.Scatter(x=complete["cycle"], y=complete["coulombic_eff_pct"], mode="markers+lines",
                              marker=dict(color=_COLORS[0]), showlegend=False), row=1, col=2)

    fig.update_xaxes(title_text="Degradation cycle #", dtick=1, row=1, col=1)
    fig.update_xaxes(title_text="Degradation cycle #", dtick=1, row=1, col=2)
    fig.update_yaxes(title_text="Ah", row=1, col=1)
    fig.update_yaxes(title_text="Coulombic efficiency [%]", row=1, col=2)
    fig.update_layout(height=440, template="plotly_white", legend=dict(orientation="h", y=-0.2))
    return fig


# ── Export ─────────────────────────────────────────────────────────────────────

def _style_ws(ws) -> None:
    """Navy bold header row + auto column widths (matches battery_feature_dashboard.py)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    for col_cells in ws.columns:
        width = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(width + 3, 42)


def export_excel(out_dir: str, dataset: pd.DataFrame, degr_df: pd.DataFrame, soh: pd.DataFrame) -> str:
    os.makedirs(out_dir, exist_ok=True)
    fname = f"{BATTERY_ID}_export_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = os.path.join(out_dir, fname)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        dataset.drop(columns=["filename"], errors="ignore").to_excel(writer, sheet_name="Run Summary", index=False)
        degr_df.to_excel(writer, sheet_name="Degradation Cycles", index=False)
        soh.to_excel(writer, sheet_name="SOH History", index=False)
        build_file_inventory(LOG_DIR).to_excel(writer, sheet_name="File Inventory", index=False)

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        if ws.max_row >= 1:
            _style_ws(ws)
    wb.save(out_path)
    return out_path


def _load_dataset() -> pd.DataFrame:
    """build_dataset() plus a derived soc_pct column (from ocv_label_V) so
    hovers/axes can show SoC alongside OCV without recomputing it everywhere."""
    df = v2feat.build_dataset(LOG_DIR, battery_id=BATTERY_ID)
    if not df.empty:
        df["soc_pct"] = df["ocv_label_V"].map(_v_to_soc)
    return df


# ── App state ──────────────────────────────────────────────────────────────────
_DATASET = _load_dataset()
_DEGR_DF = find_degr_cycles(LOG_DIR)
_SOH_DF = v2feat.find_block_soh_df(LOG_DIR, battery_id=BATTERY_ID)


def _run_options() -> dict:
    infos = [r for r in v2feat.find_v2_testday_runs(LOG_DIR) if r["ocv_label"] is not None]
    infos.sort(key=lambda r: (r["block"] or 0, r["ocv_label"]), reverse=True)
    return {
        f"Block {r['block']}  ·  OCV {r['ocv_label']:.2f}V  ·  {r['ts']}": r
        for r in infos
    }


RUN_OPTIONS = _run_options()


# ── Widgets ────────────────────────────────────────────────────────────────────
stat_blocks = pn.indicators.Number(
    name="Blocks started", value=0, format="{value}",
    default_color=_COLORS[0], font_size="32pt", title_size="12pt", align="center",
)
stat_runs = pn.indicators.Number(
    name="Test-day runs", value=0, format="{value}",
    default_color=_COLORS[2], font_size="32pt", title_size="12pt", align="center",
)
stat_cycles = pn.indicators.Number(
    name="Degr. cycles complete", value=0, format="{value}",
    default_color=_STATUS_GOOD, font_size="32pt", title_size="12pt", align="center",
)
stat_soh = pn.indicators.Number(
    name="Latest measured SOH", value=0, format="{value:.2f} Ah",
    default_color=_COLORS[0], font_size="32pt", title_size="12pt", align="center",
)
header_caption = pn.pane.Markdown("", sizing_mode="stretch_width")
rescan_btn = pn.widgets.Button(name="🔄  Rescan log folder", button_type="primary", width=200)

run_sel = pn.widgets.Select(name="Test-day run", options=list(RUN_OPTIONS), width=380)
run_pane = pn.pane.Plotly(go.Figure(), config={"responsive": True}, sizing_mode="stretch_width")
run_tbl = pn.widgets.Tabulator(pd.DataFrame(), height=400, sizing_mode="stretch_width")

compare_sel = pn.widgets.MultiSelect(
    name="Runs to compare  (Ctrl+click for multiple)",
    options=list(RUN_OPTIONS), size=10, width=420,
    value=list(RUN_OPTIONS)[:3],
)
compare_pane = pn.pane.Plotly(go.Figure(), config={"responsive": True}, sizing_mode="stretch_width")

sweep_feat_sel = pn.widgets.Select(name="Feature", options=_feature_options(_DATASET), width=320)
sweep_pane = pn.pane.Plotly(go.Figure(), config={"responsive": True}, sizing_mode="stretch_width")
sweep_tbl = pn.widgets.Tabulator(_DATASET, pagination="local", page_size=15, height=340, sizing_mode="stretch_width")

trend_feat_sel = pn.widgets.Select(name="Feature", options=_feature_options(_DATASET), width=320)
trend_pane = pn.pane.Plotly(go.Figure(), config={"responsive": True}, sizing_mode="stretch_width")

soh_pane = pn.pane.Plotly(go.Figure(), config={"responsive": True}, sizing_mode="stretch_width")

degr_pane = pn.pane.Plotly(go.Figure(), config={"responsive": True}, sizing_mode="stretch_width")
degr_tbl = pn.widgets.Tabulator(_DEGR_DF, pagination="local", page_size=15, height=340, sizing_mode="stretch_width")

inventory_tbl = pn.widgets.Tabulator(
    build_file_inventory(LOG_DIR), pagination="local", page_size=25,
    height=500, sizing_mode="stretch_width",
)

corr_clean_toggle = pn.widgets.Checkbox(name="Exclude coarse-sampled / cutoff runs (recommended)", value=True)
corr_status = pn.pane.Markdown("", sizing_mode="stretch_width")
corr_heatmap_pane = pn.pane.Plotly(go.Figure(), config={"responsive": True}, sizing_mode="stretch_width")
corr_x_sel = pn.widgets.Select(name="X axis", options=[], width=260)
corr_y_sel = pn.widgets.Select(name="Y axis", options=[], width=260)
corr_color_sel = pn.widgets.Select(name="Colour by", options=[], width=260)
corr_scatter_pane = pn.pane.Plotly(go.Figure(), config={"responsive": True}, sizing_mode="stretch_width")
predictor_tbl = pn.widgets.Tabulator(pd.DataFrame(), height=380, sizing_mode="stretch_width")
soc_robust_tbl = pn.widgets.Tabulator(pd.DataFrame(), height=380, sizing_mode="stretch_width")

export_dir_input = pn.widgets.TextInput(name="Export folder", value=os.path.join(BASE_DIR, "exports"), width=420)
export_btn = pn.widgets.Button(name="⬇  Export to Excel", button_type="primary", width=200)
export_status = pn.pane.Markdown("", sizing_mode="stretch_width")


def _update_header() -> None:
    blocks_seen = set(_DATASET["block"].unique()) if not _DATASET.empty else set()
    blocks_seen |= set(_SOH_DF["block"].unique()) if not _SOH_DF.empty else set()
    n_blocks = len(blocks_seen)
    n_runs = len(_DATASET)
    n_soh = len(_SOH_DF)
    n_soh_valid = int(_SOH_DF["reached_cutoff"].sum()) if not _SOH_DF.empty else 0
    n_cycles = len(_DEGR_DF)
    n_complete_cycles = int(_DEGR_DF["status"].str.startswith("complete").sum()) if not _DEGR_DF.empty else 0
    n_skipped = _DATASET.attrs.get("n_skipped_legacy_schema", 0)

    valid_soh = _SOH_DF[_SOH_DF["reached_cutoff"]] if not _SOH_DF.empty else _SOH_DF
    latest_ah = float(valid_soh.sort_values("block")["block_c5_capacity_ah"].iloc[-1]) if not valid_soh.empty else 0.0
    pct_nominal = latest_ah / NOMINAL_AH * 100 if latest_ah else 0.0
    soh_color = _STATUS_GOOD if pct_nominal >= 50 else (_STATUS_WARNING if pct_nominal >= 20 else _STATUS_CRITICAL)

    stat_blocks.value = n_blocks
    stat_runs.value = n_runs
    stat_cycles.value = n_complete_cycles
    stat_cycles.format = f"{{value}} / {n_cycles}"
    stat_soh.value = latest_ah
    stat_soh.default_color = soh_color

    soh_note = f"{n_soh_valid}/{n_soh} SOH measurement(s) valid" if n_soh else "no SOH measurements yet"
    header_caption.object = (
        f"**Battery: `{BATTERY_ID}`**  ·  {soh_note}"
        + (f"  ·  {n_skipped} legacy-schema file(s) skipped" if n_skipped else "")
        + "  ·  *testing in progress — click Rescan after new data lands*"
    )


_update_header()


# ── Callbacks ──────────────────────────────────────────────────────────────────

def _update_run_inspector(event=None):
    if not RUN_OPTIONS:
        return
    info = RUN_OPTIONS[run_sel.value]
    run_pane.object = build_event_fig(info["bdps_path"], info["sensor_path"])
    run_tbl.value = build_feature_table(info["bdps_path"], info["sensor_path"])


def _update_compare(event=None):
    compare_pane.object = build_compare_fig(compare_sel.value, RUN_OPTIONS)


def _update_sweep(event=None):
    sweep_pane.object = build_sweep_fig(_DATASET, sweep_feat_sel.value)


def _update_trend(event=None):
    trend_pane.object = build_trend_fig(_DATASET, trend_feat_sel.value)


def _corr_source_df() -> pd.DataFrame:
    df = _DATASET
    if corr_clean_toggle.value and "cutoff_hit" in df.columns:
        df = df[~df["cutoff_hit"]]
    return df


def _update_scatter(event=None):
    df = _corr_source_df()
    corr_scatter_pane.object = build_scatter_pair(df, corr_x_sel.value, corr_y_sel.value, color_col=corr_color_sel.value)


def _update_corr(event=None):
    df = _corr_source_df()
    cols = _corr_columns(df)
    corr_heatmap_pane.object = build_run_corr_heatmap(df, cols)
    corr_x_sel.options = cols
    corr_y_sel.options = cols
    corr_color_sel.options = cols
    if cols:
        if corr_x_sel.value not in cols:
            corr_x_sel.value = "ocv_label_V" if "ocv_label_V" in cols else cols[0]
        if corr_y_sel.value not in cols:
            r_cols = [c for c in cols if "R_int" in c]
            corr_y_sel.value = r_cols[0] if r_cols else cols[-1]
        if corr_color_sel.value not in cols:
            corr_color_sel.value = "block" if "block" in cols else cols[0]
    n_blocks = df["block"].nunique() if not df.empty else 0
    corr_status.object = (
        f"Showing **{len(df)} run(s)** across **{n_blocks} block(s)**"
        + (" (coarse-sampled/cutoff runs excluded)" if corr_clean_toggle.value else " (all runs included, including coarse-sampled/cutoff ones)")
    )
    _update_scatter()
    predictor_tbl.value = build_block_predictor_table(_DATASET, _SOH_DF)
    soc_robust_tbl.value = build_soc_robust_indicator_table(_DATASET, _SOH_DF)


def _rescan(event=None):
    global _DATASET, _DEGR_DF, _SOH_DF, RUN_OPTIONS
    _DATASET = _load_dataset()
    _DEGR_DF = find_degr_cycles(LOG_DIR)
    _SOH_DF = v2feat.find_block_soh_df(LOG_DIR, battery_id=BATTERY_ID)
    RUN_OPTIONS = _run_options()

    run_sel.options = list(RUN_OPTIONS)
    if RUN_OPTIONS and run_sel.value not in RUN_OPTIONS:
        run_sel.value = list(RUN_OPTIONS)[0]
    compare_sel.options = list(RUN_OPTIONS)

    feat_opts = _feature_options(_DATASET)
    sweep_feat_sel.options = feat_opts
    trend_feat_sel.options = feat_opts

    sweep_tbl.value = _DATASET
    degr_tbl.value = _DEGR_DF
    inventory_tbl.value = build_file_inventory(LOG_DIR)

    _update_header()
    _update_run_inspector()
    _update_compare()
    _update_sweep()
    _update_trend()
    _update_corr()
    soh_pane.object = build_soh_fig(_SOH_DF)
    degr_pane.object = build_degr_fig(_DEGR_DF)


def _run_export(event=None):
    try:
        path = export_excel(export_dir_input.value, _DATASET, _DEGR_DF, _SOH_DF)
        export_status.object = f"Saved: `{path}`"
    except Exception as e:
        export_status.object = f"Export failed: {e}"


run_sel.param.watch(_update_run_inspector, "value")
compare_sel.param.watch(_update_compare, "value")
sweep_feat_sel.param.watch(_update_sweep, "value")
trend_feat_sel.param.watch(_update_trend, "value")
corr_clean_toggle.param.watch(_update_corr, "value")
corr_x_sel.param.watch(_update_scatter, "value")
corr_y_sel.param.watch(_update_scatter, "value")
corr_color_sel.param.watch(_update_scatter, "value")
rescan_btn.on_click(_rescan)
export_btn.on_click(_run_export)

if RUN_OPTIONS:
    run_sel.value = list(RUN_OPTIONS)[0]
_update_run_inspector()
_update_compare()
_update_sweep()
_update_trend()
_update_corr()
soh_pane.object = build_soh_fig(_SOH_DF)
degr_pane.object = build_degr_fig(_DEGR_DF)


# ── Layout ─────────────────────────────────────────────────────────────────────
inspector_tab = pn.Column(
    pn.pane.Markdown("## Test Run Inspector"),
    pn.pane.Markdown(
        "Voltage/current for one test-day run, shaded by scripted event "
        "(`ocv_window → wakeup_load → glow_plug_like_load → crank_pulse → recovery_rest → "
        "… → alternator_charge → ramp_like_load / driving_aux_load`)."
    ),
    run_sel, run_pane,
    pn.pane.Markdown("### Extracted features"),
    run_tbl,
    sizing_mode="stretch_width",
)

compare_tab = pn.Column(
    pn.pane.Markdown("## Compare Runs"),
    pn.pane.Markdown("Overlay voltage/current for multiple test-day runs on a shared elapsed-time axis."),
    compare_sel, compare_pane,
    sizing_mode="stretch_width",
)

sweep_tab = pn.Column(
    pn.pane.Markdown("## SoC Sweep"),
    pn.pane.Markdown(
        "Each test-day run starts at a different OCV/SoC (set by the step-down sequence). "
        "Plotting a feature vs starting OCV shows how battery behaviour changes with SoC."
    ),
    sweep_feat_sel, sweep_pane,
    pn.pane.Markdown("### Run feature table"),
    sweep_tbl,
    sizing_mode="stretch_width",
)

trends_tab = pn.Column(
    pn.pane.Markdown("## Degradation Trends"),
    pn.pane.Markdown(
        "Any feature averaged per block and plotted vs block number. "
        "Will only show real trend structure once more than one block has been tested."
    ),
    trend_feat_sel, trend_pane,
    sizing_mode="stretch_width",
)

corr_intro = pn.pane.Alert(
    "**What this page does.** Every extracted test-day feature (crank R_int, wakeup sag, "
    "alternator charge ratio, recovery voltages, …) is correlated (Spearman) against every other "
    "feature — including starting OCV/SoC and block number — across all clean test-day runs. "
    "Below that, a focused check on the question the v2 profile was designed around: the vehicle "
    "only ever observes **OCV at rest**, never true SOH — so does any within-block feature track "
    "**this block's** or **the next block's** measured C/5 capacity ahead of time? A feature that "
    "predicts the *next* block's capacity is a candidate field-deployable health indicator.\n\n"
    "Every dot on the scatter below is one test-day run — hover it to see exactly which file, "
    "block, OCV and SoC% it is.",
    alert_type="info", sizing_mode="stretch_width",
)

corr_tab = pn.Column(
    pn.pane.Markdown("## Feature Correlations", margin=(5, 5, 0, 5)),
    corr_intro,
    pn.Card(
        pn.Row(corr_clean_toggle, sizing_mode="stretch_width"),
        corr_status,
        corr_heatmap_pane,
        title="🔗  Correlation heatmap — all features, all runs",
        collapsed=False, sizing_mode="stretch_width",
    ),
    pn.Card(
        pn.Row(corr_x_sel, corr_y_sel, corr_color_sel, sizing_mode="stretch_width"),
        corr_scatter_pane,
        title="🔍  Pairwise scatter explorer",
        collapsed=False, sizing_mode="stretch_width",
    ),
    pn.Card(
        pn.pane.Markdown(
            "One row per feature (mean over that block's clean runs) correlated against **this "
            "block's** and **the next block's** measured C/5 capacity.\n\n"
            "**Still read with some caution:** n is now ~9-10 blocks (much better than the "
            "earlier n=4, and capacity is no longer strictly monotonic — Block 6 actually came in "
            "*higher* than Block 5 — so a value of exactly ±1.0 here is a real, less-likely-by-chance "
            "signal rather than a guaranteed small-sample artifact, but it's still not a large "
            "sample. Treat rankings as directional, and expect them to keep shifting a bit as more "
            "blocks land."
        ),
        predictor_tbl,
        title="📈  Block-level SOH predictors  (this block / next block)",
        collapsed=False, sizing_mode="stretch_width",
    ),
    pn.Card(
        pn.pane.Markdown(
            "**The question that actually matters for a field-deployable indicator:** a vehicle "
            "only ever sees OCV/SoC at rest, never true SOH — so a *useful* health feature has to "
            "track degradation **independently of SoC**, not just move because the battery happens "
            "to be at a different charge level this time. `quality = |r vs next-block SOH| − |r vs "
            "SoC%|` — high when a feature strongly predicts the next block's capacity while barely "
            "moving with SoC.\n\n"
            "**Pattern so far:** resistance-derived features (crank/wakeup/ramp `R_int`) come out on "
            "top — they reflect internal impedance, not charge level, which is exactly what you'd "
            "expect physically. Absolute voltage features (`ocv_window_Vpre`, `rest_baseline_Vmean`, "
            "recovery/crank `V_pre` and `V_min`, `final_voltage`, …) rank at the bottom with "
            "`r_soc` ≈ 0.98-0.99 — they're almost fully redundant with SoC by construction, since "
            "terminal voltage directly encodes charge state for a lead-acid cell, so they add little "
            "independent health signal on their own."
        ),
        soc_robust_tbl,
        title="🎯  SoC-robust SOH indicators  (tracks health, not just charge level)",
        collapsed=False, sizing_mode="stretch_width",
    ),
    sizing_mode="stretch_width",
)

soh_tab = pn.Column(
    pn.pane.Markdown("## SOH History"),
    pn.pane.Markdown(
        "Measured C/5 discharge capacity per block — the one honest SOH ground truth "
        "(`Block_<nn>_SOH_C5_bdps_*.csv`, once per block)."
    ),
    soh_pane,
    sizing_mode="stretch_width",
)

degr_tab = pn.Column(
    pn.pane.Markdown("## Degradation Cycles"),
    pn.pane.Markdown(
        "Charge/discharge Ah and coulombic efficiency per degradation cycle within the current "
        "block (10 cycles/block). Fragment or still-in-progress cycles are marked with **✕** "
        "rather than being silently dropped."
    ),
    degr_pane,
    pn.pane.Markdown("### Cycle table"),
    degr_tbl,
    sizing_mode="stretch_width",
)

inventory_tab = pn.Column(
    pn.pane.Markdown("## File Inventory"),
    pn.pane.Markdown("One row per primary (bdps) log file found for this battery — use this to track what's landed so far."),
    inventory_tbl,
    sizing_mode="stretch_width",
)

export_tab = pn.Column(
    pn.pane.Markdown("## Export to Excel"),
    pn.pane.Markdown(
        "| Sheet | Contents |\n|---|---|\n"
        "| **Run Summary** | Feature values per test-day run |\n"
        "| **Degradation Cycles** | Ah + coulombic efficiency per cycle |\n"
        "| **SOH History** | Measured C/5 Ah capacity per block |\n"
        "| **File Inventory** | Every primary log file found on disk |"
    ),
    pn.Row(export_dir_input, sizing_mode="stretch_width"),
    export_btn, export_status,
    sizing_mode="stretch_width",
)

incident_note = pn.pane.Alert(
    "**Data notes.** ✅ A Pi-side BDPS logging-rate regression (Block 4-5, 2026-07-11 to "
    "2026-07-16) made some raw Ah numbers look far worse than reality — resolved, and SOH C/5 / "
    "Degr *discharge* Ah for the affected window are corrected here (`duration × mean current`); "
    "Degr *charge* (CC/CV) and event-level features for any still-coarse Block 5 run remain "
    "unreliable. ℹ️ A block's SOH point can also show as **\"not final yet\"** (amber ✕, excluded "
    "from trends) if that block's C/5 discharge was still running on the Pi when last synced — "
    "not an error, just rescan once it completes. See `CLAUDE.md` → \"⚠ Known Pi-side "
    "logging-rate regression\" for the full write-up.",
    alert_type="info", sizing_mode="stretch_width",
)

_STAT_CARD_STYLE = {
    "background": "#fcfcfb",
    "border": "1px solid rgba(11,11,11,0.10)",
    "border-radius": "10px",
    "padding": "6px 4px",
}


def _stat_card(indicator: pn.indicators.Number) -> pn.Column:
    return pn.Column(indicator, styles=_STAT_CARD_STYLE, sizing_mode="stretch_width")


stat_row = pn.Row(
    _stat_card(stat_blocks), _stat_card(stat_runs), _stat_card(stat_cycles), _stat_card(stat_soh),
    sizing_mode="stretch_width",
)

main_content = pn.Column(
    stat_row,
    header_caption,
    incident_note,
    pn.Row(rescan_btn),
    pn.Tabs(
        ("Run Inspector", inspector_tab),
        ("Compare Runs", compare_tab),
        ("SoC Sweep", sweep_tab),
        ("Degradation Trends", trends_tab),
        ("Feature Correlations", corr_tab),
        ("SOH History", soh_tab),
        ("Degradation Cycles", degr_tab),
        ("File Inventory", inventory_tab),
        ("Export", export_tab),
        sizing_mode="stretch_both",
    ),
    sizing_mode="stretch_both",
)

dashboard = pn.template.FastListTemplate(
    title=f"Battery Dashboard — {BATTERY_ID}",
    main=[main_content],
    accent_base_color=_COLORS[0],
    header_background=_COLORS[0],
    header_color="#FFFFFF",
)
dashboard.servable()

if __name__ == "__main__":
    pn.serve(dashboard, title=f"Battery Dashboard — {BATTERY_ID}", show=True, port=5008)
